"""
ContraXt — OpenAI-powered contract term extraction
(HLB HAMT)

Upload PDF contracts, define whatever fields you want extracted,
and get a table (+ Excel export) with the extracted values per document.

Run:
    pip install -r requirements.txt
    streamlit run app.py

API key:
    Create a file named `.env` in the same folder as this script with:
        OPENAI_API_KEY=sk-...
    Never commit this file to GitHub — keep it local only (see .gitignore).
    On Streamlit Cloud, set OPENAI_API_KEY under App settings > Secrets instead.
"""

import io
import json
import os
import pickle
import re

import openpyxl
import pandas as pd
import pdfplumber
import requests
import streamlit as st
from dotenv import load_dotenv

load_dotenv()

OPENAI_MODEL = "gpt-5.6-terra"
MAX_CHARS = 80000  # contract text sent to the model per document (~20k tokens — comfortably covers most contracts)
PERSIST_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".app_state.pkl")

# Keys that hold real data worth persisting across a page refresh. Dynamic
# per-row widget keys (overrides, reviewer picks, plan choices) are also
# persisted via the prefix list below — everything else (raw file_uploader
# widget state, nav buttons) is left alone since it can't/shouldn't be
# pre-seeded before its widget is created.
PERSIST_KEYS = [
    "fields", "results", "source_items", "pdf_bytes", "truncation_warnings",
    "confirmed_audit_df", "classification_results", "confirmed_classification",
    "confirmed_account_names", "actuals_dump_mapping", "forecast_workbook_bytes",
    "forecast_sheets_touched", "forecast_budget_signature", "wizard_step",
    "max_unlocked_step", "placement_cache", "selected_source",
    "budget_method", "fixed_budget_baseline",
    "forecast_budget_bytes", "forecast_budget_name",
    "forecast_through_month", "forecast_year",
]
PERSIST_PREFIXES = (
    "override_", "reviewer_class_", "plan_heading_", "plan_sub_",
    "plan_account_", "plan_total_", "fix_effdate_", "fix_term_", "fix_value_",
    "month_col_",
)


def save_persisted_state():
    """Best-effort — persistence failing should never break the app."""
    try:
        data = {k: st.session_state[k] for k in PERSIST_KEYS if k in st.session_state}
        for k in st.session_state.keys():
            if k.startswith(PERSIST_PREFIXES):
                data[k] = st.session_state[k]
        with open(PERSIST_FILE, "wb") as f:
            pickle.dump(data, f)
    except Exception:
        pass


def load_persisted_state():
    if not os.path.exists(PERSIST_FILE):
        return
    try:
        with open(PERSIST_FILE, "rb") as f:
            data = pickle.load(f)
        for k, v in data.items():
            st.session_state[k] = v
    except Exception:
        pass


def clear_persisted_state():
    try:
        if os.path.exists(PERSIST_FILE):
            os.remove(PERSIST_FILE)
    except Exception:
        pass
    for k in list(st.session_state.keys()):
        del st.session_state[k]


st.set_page_config(page_title="ContraXt", page_icon="📄", layout="wide")

# ---------------------------------------------------------------------------
# Styling — dark TideLedger-inspired canvas
# ---------------------------------------------------------------------------
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@1,700;1,800&family=Inter:wght@400;500;600;700&display=swap');

    html, body, [class*="css"], .stApp, [data-testid="stAppViewContainer"] {
        font-family: 'Inter', sans-serif;
        color: #F4FBFF;
    }
    .stApp {
        background:
            radial-gradient(1100px 420px at 12% -8%, rgba(46, 200, 222, 0.10), transparent 55%),
            linear-gradient(180deg, #07161f 0%, #06141c 100%);
    }
    [data-testid="stHeader"] { background: transparent; }
    [data-testid="stToolbar"] { background: transparent; }
    [data-testid="stSidebar"] {
        background: #0a1c26;
        border-right: 1px solid rgba(94, 210, 230, 0.18);
    }
    [data-testid="stSidebar"] * { color: #E8F4F8; }
    .stMarkdown, .stCaption, label, p, span, div { color: inherit; }
    [data-testid="stCaption"] { color: #8AA3B0 !important; }

    .block-container { padding-top: 1.1rem !important; padding-bottom: 2rem !important; }
    [data-testid="stSidebar"] img { max-height: 44px !important; width: auto !important; }
    .cx-header {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 14px;
        margin: 0 0 10px;
    }
    .cx-logo {
        background: #F4FBFF;
        border-radius: 8px;
        padding: 5px 8px;
        height: 48px;
        display: flex;
        align-items: center;
        flex-shrink: 0;
    }
    .cx-logo img { height: 36px; width: auto; display: block; }
    .wordmark-wrap { text-align: left; margin: 0; }
    .wordmark {
        margin: 0;
        font-family: "Playfair Display", "Times New Roman", serif;
        font-style: italic;
        font-weight: 800;
        font-size: 44px;
        letter-spacing: 0.03em;
        line-height: 1;
        color: #F3FBFF;
        -webkit-text-stroke: 1px #16323c;
        text-shadow:
            0 1px 0 #ffffff,
            0 0 3px rgba(190, 255, 230, 0.9),
            0 0 10px rgba(94, 228, 242, 0.55),
            0 0 18px rgba(46, 200, 222, 0.25),
            2px 3px 0 #0a2430;
    }
    .wordmark-sub {
        margin-top: 5px;
        color: #8AA3B0;
        font-size: 0.78rem;
        letter-spacing: 0.06em;
        text-transform: none;
    }
    .app-title { display: none; }
    .app-subtitle { display: none; }

    .step-label {
        font-family: 'Inter', sans-serif;
        font-weight: 600;
        font-size: 1.05rem;
        color: #F4FBFF;
        border-bottom: 1px solid rgba(94, 210, 230, 0.22);
        padding-bottom: 0.4rem;
        margin-top: 1.6rem;
        margin-bottom: 0.9rem;
    }
    div[data-testid="stStatusWidget"] { display: none; }
    #MainMenu, footer { visibility: hidden; }

    div.stButton > button {
        background: #0b2130 !important;
        color: #F4FBFF !important;
        border: 1px solid rgba(94, 228, 242, 0.35) !important;
        border-radius: 999px !important;
        box-shadow: 0 0 0 1px rgba(94, 228, 242, 0.16), 0 0 14px rgba(46, 200, 222, 0.12);
    }
    div.stButton > button:hover {
        border-color: #5EE4F2 !important;
        box-shadow: 0 0 0 1px rgba(94, 228, 242, 0.45), 0 0 18px rgba(46, 200, 222, 0.28) !important;
    }
    div.stButton > button[kind="primary"] {
        background: #0b2834 !important;
        color: #F4FBFF !important;
        border: 1px solid #5EE4F2 !important;
        box-shadow: 0 0 0 1px rgba(94, 228, 242, 0.45), 0 0 22px rgba(46, 200, 222, 0.38) !important;
    }
    div.stButton > button[kind="primary"]:hover {
        background: #10303c !important;
    }
    div.stButton > button:disabled {
        opacity: 0.42;
        box-shadow: none !important;
        border-color: rgba(138, 163, 176, 0.25) !important;
        color: #8AA3B0 !important;
    }

    [data-testid="stTextInput"] input,
    [data-testid="stNumberInput"] input,
    [data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    textarea {
        background: #0a1c26 !important;
        color: #F4FBFF !important;
        border-color: rgba(94, 210, 230, 0.28) !important;
    }
    [data-testid="stExpander"] {
        background: #0c2230;
        border: 1px solid rgba(94, 210, 230, 0.18);
        border-radius: 14px;
    }
    [data-testid="stFileUploader"] {
        background: #0c2230;
        border: 1.5px dashed rgba(94, 228, 242, 0.28);
        border-radius: 16px;
        padding: 0.4rem;
    }
    table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.9rem;
        color: #F4FBFF;
    }
    table th {
        background-color: #0e2836;
        color: #5EE4F2;
        text-align: left;
        padding: 0.5rem 0.7rem;
        border-bottom: 1px solid rgba(94, 210, 230, 0.28);
    }
    table td {
        padding: 0.5rem 0.7rem;
        border-bottom: 1px solid rgba(94, 210, 230, 0.14);
        color: #E8F4F8;
    }
    table tr:hover td {
        background-color: rgba(94, 228, 242, 0.06);
    }
    .logo-chip {
        display: inline-flex;
        align-items: center;
        background: #F4FBFF;
        border-radius: 10px;
        padding: 6px 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Session state
# ---------------------------------------------------------------------------
if "_state_loaded" not in st.session_state:
    load_persisted_state()
    st.session_state["_state_loaded"] = True

if "fields" not in st.session_state:
    st.session_state.fields = [
        {"name": "Counterparty", "hint": "The other party to the contract (not our company)"},
        {"name": "Effective Date", "hint": "Date the contract starts / becomes effective"},
        {"name": "Term / Expiry", "hint": "Contract duration, end date, or renewal terms"},
        {"name": "Contract Value", "hint": "Total or annual fee/value, with currency"},
    ]

if "results" not in st.session_state:
    st.session_state.results = None

if "wizard_step" not in st.session_state:
    st.session_state.wizard_step = 1
if "max_unlocked_step" not in st.session_state:
    st.session_state.max_unlocked_step = 1


# ---------------------------------------------------------------------------
# API key resolution — env var (.env) first, then Streamlit Secrets
# ---------------------------------------------------------------------------
def get_secret(key: str):
    try:
        return st.secrets[key]
    except (FileNotFoundError, KeyError, st.errors.StreamlitAPIException):
        return None


OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY") or get_secret("OPENAI_API_KEY")

# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
import base64 as _base64
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "hlb_hamt_logo.png")
_LOGO_B64 = (
    "iVBORw0KGgoAAAANSUhEUgAAANgAAAB6CAYAAADd9J0IAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8"
    "YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAABFXSURBVHhe7Z1BaBtZmsf/2Z0slkITCaKhI5iR4qV3284h"
    "ytgwWexxZOKDGjy0E8cwQ9oT+xC7Tx15DtPBGGyDaZI5tO0wF7sPdnAHFhInbjC0Dw5W3DHMwZpWH2z3"
    "YXGkXlAWFJCWJVJgBrwH1auoXr1XqpJUlmR/PygSv1KVnqrev773vve9r04cHBwcgCAIW/gnvoAgiMpB"
    "AiMIGyGBEYSNkMAIwkZIYARhIyQwgrAREhhB2AgJjCBshARGEDZCAiMIGyGBEYSNkMAIwkZIYARhIyQw"
    "grAREhhB2EhdCCyyt494Ks0XE0TNc6LaCy4z2RxiiVeI7O7jh59eIfMmh8jePv8xHQGfFy5nAy43NcLv"
    "cSOo/EsQtURVBBZLJLGyvYtvoruIJZL87pJhQvu4pRk9ref53QRx6ByawDLZHBY3o5hd2zqU7p7L6UBP"
    "azNuh9oQ8Hn53QRxKNgusHgqjckn61jcjPK7Do1gUyPGe7sQbGrkdxGErdgmsFoQFg8JjThsbBHYxPI6"
    "Zte2kMnm+F1FKXReyLDiDBHR03oe0590k1OEsJ2KCiyWSGJw7rElx0WwqREftzYj4POWZFniqTQie/t4"
    "vrePle1d06J2OR0Y772CcKid30UQFaNiAptZe4GRpVW+WEjA58XtUBt6Wpvhcjr43WWxsr2Db6K7prum"
    "Pa3nsTB8veL1IAhUQmCZbA4jS6umGvRAR8uhefUy2Rxmvt0y1VUN+LxYGL5+KPUijhdlCSyTzaFz6qui"
    "XcJgUyMWhvuqMuYxKzSX04GNsVskMqKilCwwM+Lye9xYGO4raWxVaeKpNAbnHhV1jCwM92Ggo4UvJoiS"
    "KElgZsQ10NGC6f7umhvbzKy9wOTyM0NrRiIjKoVlgRUTl8vpwHR/d003UDPezo2xoZqwvER9Y1lgF0fv"
    "SxtmPY1jMtkcrn65JO0y1tNvIWoXS8tVBuceHQlxQa3vkNTSZrI5XJ1eMuxKEkQxTAtscTMqdcXXm7gK"
    "MRpv5R0jj/ligjCNKYHFU2npJHI9i4th5Olc2d7BzNoLvpggTGFKYINzj6RdpXoXF+PpH/ulv2Ny+dmh"
    "LLEhjh5FBba4GZU6Aqb7u6WNst5wOR3SkKlMNoeRr8UWnCCMMBQYC4MS0dN6/sgFygZ8Xkz3d/PFgNJV"
    "lD1oCEKGocBmvhWHF7Gn/VFkoKNFmm5A9rAhCBlSgWWyOcyubfHFgNI1FHWljgrTn4h/XyyRlHpSCUKE"
    "VGAy6xXweaVu7aOC3+PG7VAbXwwAeEACIywgFdiD78QNSTZGOWqEP2oTWrHI3j6NxQjTCAW2sr0jdEsH"
    "mxql80VHDZfTQVaMKBuhwB589ze+CABw84h3DXlkVmxle5cvIgghOoFlsjmsbO/wxXA5HUd+7MXjUnIr"
    "8siuEUHw6KLpV7Z3cHV6qbAIABAOtR+b8Vchkb19dE7N88UY6GjBwnAfX6wST6URf63tZhfrXmeUNOKF"
    "BHxnhVZUxsr2juYcrlMNluYrRXVwORtMBxSYOV50baziP+Ouygp5q+gENjj3SOiK/v6Lz3QXeWJ5HZNP"
    "1jVlxdZRGR0jaszj17ow0dulKStGJpvD4NxjqZVhlslsfhD3rUmdR9XldCD91bimrBDR7zx4eFfzN4/o"
    "9xe7njznwvd04+eXM5+bboyiOkBy/0WMLK3qYjeDTY3YGBtS/xZdG6uU0i6qga6LKPKQuZwOUxe3Fogl"
    "kuic+koqLigCXNyM4uLofeHDhEfWTZQt3akWi5tRnbgAlN2YYXKSPbK3rxPXcUcjsEw2J7xBogZWiyxu"
    "Rg1XW4swWuPGkCVB5btC1Ubm3bSSL1KGGfGYEeFxQyMwWYO54DvLF9Uci5tRw6h/I2QRKwxZF+0HyfWq"
    "BrFEUtj7QIHFLhejXCYTy+tFH1QM9hYcfhP1kgJKQlp+M9vlrTYagUV2xTdI9MNrDd7yBnxebIwN6bbx"
    "a/p+u+zBwvB73EJHg9kGdRgUe0gU228GNrblYe8hMMtAR4vuvmyMDQmdaNP93brPbRisRK81dGMwEbIn"
    "eC3jcjbonnrBpkZM9Hbpbo4ZoQQEVrxcT1ilyE8baOfm+IBllmK8XFa2d3Tj28G5R5q/iXdoBPZccANE"
    "T+56h/9NZiy0qEvCW81qsbgZ1XTd2GoHvs6VsGIAMPL1qvp9RusFCRMWTPTkrmdEEfFmfqPvjF5gtQIv"
    "HJbz/+ZvtJZaFgJXDL4HE0+l1fEY79ioJ4/zYVBUYFbpnJrHiRt3pJuVvno5xBKv0Dk1r9kujt7HxdH7"
    "uqe9aFxmFisNlq8Pv/GN1Qwi0bAYyvBH+lhKXoxmGO/t0olmZu0Frn6pz7qVXxXeoCk7zlRcYLVCJpt/"
    "f1jhJhprTfeX954wK+Mwvj78JqpfMfi40YDPq4pBFN7GW2+ziBbY8l1Do8Wqx5UjKzCzDM49wrnwPV1j"
    "qQfiqbTO4cCvAOADtEt12Qd8XkNL7/e4hV7A486xFxiUhto5NV93IuO7e6LgZNGcEX+cWSYEXUXGwnCf"
    "znlE2CCwgY4WjF/rkm78gNlORHMoT0f6pfUoxd0sOo8Mvi78ZsUCiCxRJpuD+9akbtzLj9GMJqWLIeoq"
    "hkPtlq7DcaLiArvZ0YKJ3nwgpmiThR3ZgSgKoKf1PCZ6u7AxNqQbL1RqrkgGXxd+k1kHEeWGP8nCqorB"
    "dxX9HjfGe69oPkO8o6jA7Gxw1eZy0zm+SBrN8sNPxtEeh0253lhZYLAZCruK1DU0RiMwmXUp50lZyzzf"
    "e8kXScm80V8DKxankkT29ksWRyF8F9MKC8PXqWtoAo3AXKfE8xfFYvXs5PnePiaW16WbkYVlYw1+m1l7"
    "gc6peZ0HDgaTzqLvqdZ8D9+9czkdOHh4t+jGPxBkiY3MEDBI0kq8QyMw/gYwZN2mSuD3uOE3iJKI7O1j"
    "8kl+gZ5oM6rbyNKqbjKXTeiKBeNAsFn/RJbNT8ksvp3EU2md5eHnumTwLnzRuYjKohGYzNzbNf4INjXi"
    "+y8+07mRq4UsoapIjDCwdnYiEgQvHBkshKoQ3hoSlUXn5BBZMSMrUSrhUDs2xoZ0N7wa5INj5e8Jk43V"
    "RNbObvg5rKBgnkuGaJ4sUmIECWEOncBEVoyFHVUC16kGPB3pr3r/PeDzoqf1PBaG+/By9k9SceWXgojG"
    "al7Dh4NoUWExREtsCsd5kb19BHxnNfvHLealuB1q031H4b0tVgcz8NMjooe2iEp8d62hS3pDaGErpXmO"
    "a5Ytwho6C0ZokY1Rbnb8ii8iCB0kMANkkR1+j9t0t4c43vzzxMTEBF9I5Bn5elU4B3g71G7JwRFLJPE/"
    "//t/eN/1nlrGhOs6lR/HFX4mlkjix1cpxF+nkcnmNMfFU2nEfnoFv8eNyN4+4q/zSTzZVnhOxokbd3Q5"
    "BIsdy9ePfa/rVAMaTp5UP8PXD8q6N79HnhiUPzcrE9Wj8Pcy+Gsg+nxhPRmZbA5//a//1u3jr4Xf41Y/"
    "C5PnlkFjMAnxVBrnwvf4YsBiIs9MNodzt/8MAHg5+yfVMXLixh1N8sxOJdnnxtiQLrI/4PNiYfg6Aj6v"
    "mrTz4OFdnLhxR/0MQ5SQ88SNO7qkp0bHRpTko8GChKGZbA4XR+/Df8aNjbEhzKy9wMjSqjAhaefUPMZ7"
    "xQHVonNDUJ+8A+o64qk0Lo7e13h5L47eh9/jxtORfs11ZHViTPd3q1mN+YSohWNo/rv9Hje+/+IzXP1y"
    "CfHXabyc+RxQvtflbNDUuxjURZQgeyfzQEeLaXGh4EUReW+k+ZdGBJsacfDwrpo9WJTNiUVoQBHHwcO7"
    "OnHJMDr2wWYULqcj/2RXQrLYVEZEiYSZXH6G8Wvy5SsyROdmsHp8/8VniOzuY3L5GQLKAtJvovlrF0+l"
    "EUskcfM3+jHw5PIzhEPtOHh4Nx/M3ZIP5p5YXsfM2gssDPfh4OFdLAz36cRY+N1sAn66vxvxVFo9PpZI"
    "WnZskcAERPb2ha55KDfCCrNrW+hpbcZAR4tuDssMLqcD49euIJZIHsp8FVsGw1Z6F9Y52NSIcKgdI0ur"
    "8HvcpsXMMDp3IQHlJY9sUv12qA0r2zvIKG9d9XvcupUQ+ePO5lMZTC8hnsp39aDcg3CoXbWAAx0tCIfa"
    "hZP27JjMm7cIKCsHZte2Sn6gkMA4Mtmc0C2PEqxXRJnEveA7C98Zd8nrsNgYIJN9y++qODPf5hu9X3Hk"
    "LHIZq04r81KlzE8VO3chp50N6j42Ob64GcVKdEe1TDxP/9iPcKgdsUQSg3OPVAuVyebUejMKzw8lLrNT"
    "ydviKki1EP6oDa5TDfmXaAhynBSDBMYxufxM13VhWLVezMU/srSqLi9hZQGfVw1By2RziL9OS5+Os2tb"
    "+ThJwZim0rAAYBYMXdi1jSWSmHyyjnCoXe0qWsHo3IVksjk8+C6qWinW4Nm9kYWGxRKvMN57BS9nPkc4"
    "1K7Wr6f1PGbXttT7Gk+lMbu2pQku8J/Jiz6eSmtS3rmcjvybXM6Ik88WoyyByRpivbK4GZU2mvFrXZas"
    "F+sOsX4/6/uzp/b4tStY2d55l+nqzVtNw2FZsS6O3kdkd1+4krjSsAxVL2c+V+tc2LUdnHuMYFOj6jww"
    "ehjxFDs3OCsC5WX0jJsdLchkcwhKQsMy2RyufrmEi6P30Tk1j8XNqCqg6U/yXdJz4XvonJrHufA9+Lkc"
    "IpeV39XTeh6Dc4+lltUqJbvpJ5bXcXV6STX39U4skcTv//KfePv3f/C74Pe4sfhpn2nXLAD8mEzh/dPv"
    "YeByi3qc3+NGw8mTeN/1HoLN/4qBjhY0/MvP8LtLFzDzh26Nu/tDrwd+jxuXPvgFZvp/i0sf/DK/40T+"
    "actPEwSbxQ0PyuJMo/ESO/bHZAqXPvglQhf+Xd2nnvME0HDyZ5jo7YLrlAOXPviF+pkPvT9X/w/FSvP1"
    "MTp3wH8WDSdPIuDzwq+Mr2b6tdeDfXbgsr6bHmxuxIfen+PTrl/j/dPvweV04Hf/cUH9za5TDnx65dcI"
    "+PLfEw61Yab/t5r7yeobuvBvAIC3f/+H5ncxZ4tVLLvp82MU7bu3Niy+w6rWYK5g2VNL5IquJzqn5i25"
    "lstlZGkVNzt+VdfXrFJYElgskVQ9NIW4nA5sjN2qywuayeYMX3kkmlciCLNYGoOJ5i5Q0EhL8ZBVk5jy"
    "sj6ZuFiCHIIoFUsWDAavmGUYrauqJZi4ZN3CgM+LjbFbJXmOCIJhWWAAcHV6SToRizpYyiFbgsLwK6Ey"
    "JC6iXEoSWLFxC7j4uVpB5KDhqefxJFF7lCQwmGysUJwE4Y/aqm4NFjejGFl6914rEbX4UCDqm5IFxig2"
    "JgPL/npN/2bJwyCyt4/JIundQGMuwibKFhiUd0WZebcVE5oou1GlWdnewezaVlFhQYkxlGWUIohyqIjA"
    "oFiKwblHQjc+j0vJbvRxS7MwKrpUYokkHmz+DStR/UvpRLicDoz3XlHXDBFEpamYwGBhXFaIS0n2ebnp"
    "nBKOcta0JWHzcs+VjL1mRMVgMXU03iLspKICY1ixZjKMQq9iiVeGzgojyGoRh4ktAmMsbkYx+WS9LKFV"
    "CpfTgduhtprwaBLHB1sFxqim0EhYRDU5FIExInv7eLAZLfvlcWboaT2Pj1vyS/UJolocqsAKWdnewfO9"
    "l+qy+nLxK6mqLzc1Hso0AEGYoWoCKySTzSGWeIVYIonMm7d4XjB3FX+dVruWhY6PgM+L086G/EK5M/Ic"
    "fARRTWpCYARxVLG0HowgCGuQwAjCRkhgBGEjJDCCsBESGEHYCAmMIGyEBEYQNkICIwgbIYERhI2QwAjC"
    "RkhgBGEjJDCCsBESGEHYCAmMIGyEBEYQNkICIwgbIYERhI2QwAjCRkhgBGEjJDCCsBESGEHYCAmMIGyE"
    "BEYQNkICIwgbIYERhI2QwAjCRkhgBGEj/w/JLY2LLqscTAAAAABJRU5ErkJggg=="
)

def _load_logo():
    if os.path.exists(LOGO_PATH):
        return LOGO_PATH
    try:
        return io.BytesIO(_base64.b64decode(_LOGO_B64))
    except Exception:
        return None

_logo = _load_logo()
st.markdown(
    f"""
    <div class="cx-header">
      <div class="cx-logo"><img alt="HLB HAMT" src="data:image/png;base64,{_LOGO_B64}" /></div>
      <div class="wordmark-wrap">
        <p class="wordmark">ContraXt</p>
        <div class="wordmark-sub">AI Driven Contractual Terms Extraction &amp; Forecasting</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

if not OPENAI_API_KEY:
    st.error(
        "No OpenAI API key configured. Add OPENAI_API_KEY to a local `.env` file, "
        "or under App settings \u2192 Secrets if this is deployed on Streamlit Cloud."
    )

with st.sidebar:
    if _logo is not None:
        st.image(_logo, use_container_width=True)
        st.divider()
    st.caption("STATUS")
    if OPENAI_API_KEY:
        st.success("Connected")
    else:
        st.warning("Not connected")
    st.divider()
    st.caption(
        "Your progress is saved automatically and survives page refreshes. "
        "Contract text is sent to OpenAI's API for extraction."
    )
    st.divider()
    if st.button("🗑️ Clear everything", use_container_width=True):
        st.session_state["_confirm_clear"] = True

    if st.session_state.get("_confirm_clear"):
        st.warning("This permanently deletes all extracted, classified, and forecasted data. This can't be undone.")
        cc1, cc2 = st.columns(2)
        if cc1.button("Yes, clear it all", type="primary", use_container_width=True):
            clear_persisted_state()
            st.rerun()
        if cc2.button("Cancel", use_container_width=True):
            st.session_state["_confirm_clear"] = False
            st.rerun()


# ---------------------------------------------------------------------------
# Step 1: define fields
# ---------------------------------------------------------------------------
def extract_pdf_text(file) -> str:
    text_parts = []
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            text_parts.append(page_text)
    return "\n".join(text_parts)


def apply_reviewer_overrides(results_df: pd.DataFrame) -> pd.DataFrame:
    """Applies any reviewer edits stored under override_{row}_{col} keys to a copy
    of the audit results — used both for the on-screen table and anywhere else
    (like Forecasting) that needs the reviewer's corrected values, not the raw model output."""
    display_df = results_df.copy()
    editable_cols = [c for c in display_df.columns if c not in ("File", "Error")]
    for row_idx in display_df.index:
        for col in editable_cols:
            override_key = f"override_{row_idx}_{col}"
            if override_key in st.session_state:
                display_df.at[row_idx, col] = st.session_state[override_key]
    return display_df


def call_openai(prompt: str, api_key: str, model: str) -> str:
    resp = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
        },
        timeout=180,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def parse_json_response(raw: str) -> dict:
    cleaned = re.sub(r"^```(json)?|```$", "", raw.strip(), flags=re.MULTILINE).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", cleaned, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(0))
            except json.JSONDecodeError:
                pass
        return {"error": "Could not parse model response", "raw_response": raw[:500]}


def render_audit_mode():
    st.markdown('<div class="step-label">1 · Define the fields to extract</div>', unsafe_allow_html=True)

    for i, field in enumerate(st.session_state.fields):
        c1, c2, c3 = st.columns([2, 5, 1])
        field["name"] = c1.text_input("Field name", field["name"], key=f"name_{i}", label_visibility="collapsed", placeholder="Field name")
        field["hint"] = c2.text_input("What to look for", field["hint"], key=f"hint_{i}", label_visibility="collapsed", placeholder="What to look for")
        if c3.button("Remove", key=f"remove_{i}"):
            st.session_state.fields.pop(i)
            st.rerun()

    if st.button("+ Add field"):
        st.session_state.fields.append({"name": "", "hint": ""})
        st.rerun()


    # ---------------------------------------------------------------------------
    # Step 2: upload PDFs
    # ---------------------------------------------------------------------------
    st.markdown('<div class="step-label">2 · Upload contracts</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader("PDF files", type=["pdf"], accept_multiple_files=True, label_visibility="collapsed")


    # ---------------------------------------------------------------------------
    # Extraction helpers
    # ---------------------------------------------------------------------------
    def extract_pdf_text(file) -> str:
        text_parts = []
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                text_parts.append(page_text)
        return "\n".join(text_parts)


    def _boxes_from_chars(chars):
        """Groups matched characters by visual line, returning one bbox per line
        so a match spanning a line-wrap highlights correctly instead of one box
        stretching across both lines."""
        lines = {}
        for c in chars:
            key = round(c["top"], 1)
            lines.setdefault(key, []).append(c)
        boxes = []
        for cs in lines.values():
            boxes.append((
                min(c["x0"] for c in cs),
                min(c["top"] for c in cs),
                max(c["x1"] for c in cs),
                max(c["bottom"] for c in cs),
            ))
        return boxes


    def locate_quote(pdf_bytes: bytes, quote: str):
        """Search every page for the quote. Returns (page_number, [bbox, ...]) or None if not found.
        Multiple boxes are returned when the match spans a line wrap."""
        if not quote or not quote.strip():
            return None
        quote = quote.strip()
        words = quote.split()

        # Build a whitespace-tolerant regex so line wraps, extra spaces, or
        # non-breaking spaces in the PDF's text don't break an otherwise-correct match.
        def flexible_pattern(word_list):
            return r"\s+".join(re.escape(w) for w in word_list)

        # Try the full quote first, then progressively shorter windows (in case the
        # model's quote includes a word or two not present verbatim in the source).
        attempts = [words]
        if len(words) > 4:
            attempts.append(words[: len(words) - 2])
            attempts.append(words[2:])
        if len(words) > 6:
            attempts.append(words[2:-2])

        try:
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    for word_set in attempts:
                        if not word_set:
                            continue
                        pattern = flexible_pattern(word_set)
                        try:
                            matches = page.search(pattern, regex=True, case=False)
                        except Exception:
                            matches = []
                        if matches:
                            boxes = _boxes_from_chars(matches[0]["chars"])
                            if boxes:
                                return page_num, boxes
        except Exception:
            return None
        return None


    def render_highlighted_page(pdf_bytes: bytes, page_number: int, boxes) -> bytes:
        """Renders the given page as a PNG with each bbox in `boxes` highlighted."""
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page = pdf.pages[page_number]
            im = page.to_image(resolution=150)
            for bbox in boxes or []:
                im.draw_rect(bbox, fill=(255, 224, 102, 90), stroke=(230, 126, 14), stroke_width=2)
            buf = io.BytesIO()
            im.save(buf, format="PNG")
            return buf.getvalue()


    def build_prompt(contract_text: str, fields: list) -> str:
        field_lines = ",\n".join(
            f'    "{f["name"]}": {{"value": "<value — {f["hint"]}>", "source": "<short verbatim quote from the contract text that this value came from, 4-12 words, EXACT wording, or empty string if not found>"}}'
            for f in fields if f["name"].strip()
        )
        return f"""You are extracting key terms from a contract. Read the contract text below and
    return ONLY a JSON object with exactly this shape:

    {{
      "fields": {{
    {field_lines}
      }},
      "pricing_type": "Fixed" | "Variable" | "Mixed" | "Not found",
      "pricing_notes": "any caps, minimum fees, or rate-review terms that affect the total (short phrase), or empty string if none",
      "channels": [
        {{
          "name": "the channel, service line, or unit this rate applies to (e.g. 'Paid Search (Search Ads)')",
          "name_source": "<short verbatim quote naming this channel, EXACT wording>",
          "billing_unit": "how it's billed (e.g. 'Per Click (CPC)')",
          "rate_value": <number only, e.g. 2.75>,
          "rate_currency": "currency code or symbol, e.g. AED",
          "rate_unit_label": "what the rate is per, e.g. 'click' or '1,000 impressions'",
          "rate_per_n_units": <if the rate is per N units like 'per 1,000 impressions' put N here, else 1>,
          "rate_source": "<short verbatim quote containing this rate, EXACT wording>",
          "volume_low": <lowest estimated volume as a plain number, e.g. 6000>,
          "volume_high": <highest estimated volume as a plain number; same as volume_low if only one figure is given>,
          "volume_source": "<short verbatim quote containing this volume figure, EXACT wording>"
        }}
      ]
    }}

    Rules:
    - "fields" must contain exactly the keys listed above, each an object with "value" and "source". If a value isn't found, use "Not found" for value and "" for source — never guess or invent a value. Keep each value short (a phrase, date, or figure), not a full sentence.
    - Every "source" quote must be copied EXACTLY from the contract text below — same words, same characters, same punctuation and capitalization — so it can be located verbatim on the page. Do not paraphrase or summarize it. Keep each quote short (under ~12 words) and specific enough to be unique in the document.
    - Classify "pricing_type" as "Variable" if fees depend on usage/volume (e.g. per-click, per-impression, per-lead, commission-based), "Fixed" if it's a flat recurring or one-time fee, "Mixed" if both appear, "Not found" if unclear.
    - Populate "channels" ONLY when the contract gives per-unit rates with an associated volume or volume estimate (e.g. a rate card, media plan, or pricing table). Leave it as an empty list if the contract is a flat/fixed fee with no such breakdown.
    - Every number in "channels" must be a plain numeric value with no currency symbols, commas, or text — put units and currency in the separate label fields.
    - Do not compute or total anything yourself — just extract the raw rate and volume figures per channel; the values will be calculated separately.
    - Return ONLY the JSON object, no other text, no markdown fences.

    CONTRACT TEXT:
    \"\"\"
    {contract_text}
    \"\"\"
    """



    def call_openai(prompt: str, api_key: str, model: str) -> str:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json={
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
            },
            timeout=180,
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]


    def parse_json_response(raw: str) -> dict:
        cleaned = re.sub(r"^```(json)?|```$", "", raw.strip(), flags=re.MULTILINE).strip()
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", cleaned, re.DOTALL)
            if match:
                try:
                    return json.loads(match.group(0))
                except json.JSONDecodeError:
                    pass
            return {"error": "Could not parse model response", "raw_response": raw[:500]}


    def compute_channel_value(channel: dict):
        """Returns (estimated_value_number, display_string) or (None, 'n/a') if it can't be computed."""
        try:
            rate = float(channel.get("rate_value"))
            vol_low = float(channel.get("volume_low"))
            vol_high = float(channel.get("volume_high"))
            per_n = float(channel.get("rate_per_n_units") or 1) or 1
        except (TypeError, ValueError):
            return None, "n/a"

        mid_volume = (vol_low + vol_high) / 2
        value = rate * (mid_volume / per_n)
        currency = channel.get("rate_currency") or ""
        display = f"{currency} {value:,.0f}".strip()
        return value, display


    def build_result_rows(file_name: str, parsed: dict, active_fields: list) -> list:
        if "error" in parsed:
            return [{"File": file_name, **parsed}]

        field_values = parsed.get("fields", {})
        base = {"File": file_name}
        for f in active_fields:
            entry = field_values.get(f["name"], {})
            value = entry.get("value", "Not found") if isinstance(entry, dict) else entry
            base[f["name"]] = value
        base["Pricing Type"] = parsed.get("pricing_type", "Not found")

        channels = parsed.get("channels") or []
        if not channels:
            base["Notes"] = parsed.get("pricing_notes", "")
            return [base]

        rows = []
        total_value = 0.0
        any_computed = False
        total_currency = ""
        for ch in channels:
            value, display = compute_channel_value(ch)
            if value is not None:
                total_value += value
                any_computed = True
                total_currency = total_currency or (ch.get("rate_currency") or "")
            rate_str = f"{ch.get('rate_currency', '')} {ch.get('rate_value', '')} / {ch.get('rate_unit_label', '')}".strip()
            vol_low, vol_high = ch.get("volume_low"), ch.get("volume_high")
            vol_str = f"{vol_low:,}".rstrip() if vol_low == vol_high else f"{vol_low:,} – {vol_high:,}" if vol_low is not None and vol_high is not None else "Not found"
            row = dict(base)
            row.update({
                "Channel": ch.get("name", "Not found"),
                "Billing Unit": ch.get("billing_unit", "Not found"),
                "Rate": rate_str,
                "Est. Volume": vol_str,
                "Est. Value": display,
            })
            rows.append(row)

        if any_computed:
            total_row = dict(base)
            notes = parsed.get("pricing_notes", "")
            total_row.update({
                "Channel": "TOTAL (all channels)",
                "Billing Unit": "",
                "Rate": "",
                "Est. Volume": "",
                "Est. Value": f"~ {total_currency} {total_value:,.0f}".strip(),
            })
            if notes:
                total_row["Est. Value"] += f"  (note: {notes})"
            rows.append(total_row)

        return rows


    def build_source_items(file_name: str, parsed: dict, active_fields: list) -> list:
        """Flat list of {label, value, quote} for the click-to-verify panel."""
        if "error" in parsed:
            return []

        items = []
        field_values = parsed.get("fields", {})
        for f in active_fields:
            entry = field_values.get(f["name"], {})
            if isinstance(entry, dict):
                value, quote = entry.get("value", "Not found"), entry.get("source", "")
            else:
                value, quote = entry, ""
            items.append({"label": f["name"], "value": value, "quote": quote})

        for i, ch in enumerate(parsed.get("channels") or [], start=1):
            name = ch.get("name", f"Channel {i}")
            rate_display = f"{ch.get('rate_currency', '')} {ch.get('rate_value', '')} / {ch.get('rate_unit_label', '')}".strip()
            vol_display = f"{ch.get('volume_low', '')} – {ch.get('volume_high', '')}"
            items.append({"label": f"{name} — Name", "value": name, "quote": ch.get("name_source", "")})
            items.append({"label": f"{name} — Rate", "value": rate_display, "quote": ch.get("rate_source", "")})
            items.append({"label": f"{name} — Volume", "value": vol_display, "quote": ch.get("volume_source", "")})

        return items


    # ---------------------------------------------------------------------------
    # Step 3: run extraction
    # ---------------------------------------------------------------------------
    st.markdown('<div class="step-label">3 · Extract</div>', unsafe_allow_html=True)

    active_fields = [f for f in st.session_state.fields if f["name"].strip()]

    if "source_items" not in st.session_state:
        st.session_state.source_items = {}  # file_name -> list of {label, value, quote}
    if "pdf_bytes" not in st.session_state:
        st.session_state.pdf_bytes = {}  # file_name -> raw bytes
    if "selected_source" not in st.session_state:
        st.session_state.selected_source = None  # (file_name, item_index)

    if st.button(
        "Extract key terms",
        type="primary",
        disabled=not (uploaded_files and active_fields and OPENAI_API_KEY),
    ):
        # Clear any stale manual overrides from a previous extraction run —
        # row indices reset each run, so leftover overrides could apply to the wrong row.
        for k in list(st.session_state.keys()):
            if k.startswith("override_"):
                del st.session_state[k]
        # A fresh extraction invalidates any downstream confirmed snapshots —
        # force re-confirmation through Classification and Forecasting.
        st.session_state.pop("confirmed_audit_df", None)
        st.session_state.pop("confirmed_classification", None)
        st.session_state.max_unlocked_step = 1

        rows = []
        source_items = {}
        pdf_bytes_map = {}
        truncation_warnings = []
        progress = st.progress(0.0, text="Starting...")

        for idx, file in enumerate(uploaded_files):
            progress.progress(idx / len(uploaded_files), text=f"Reading {file.name}...")
            pdf_bytes_map[file.name] = file.getvalue()
            try:
                text = extract_pdf_text(file)
            except Exception as e:
                rows.append({"File": file.name, "Error": f"Failed to read PDF: {e}"})
                continue

            if not text.strip():
                rows.append({"File": file.name, "Error": "No extractable text (likely a scanned/image PDF — needs OCR)"})
                continue

            if len(text) > MAX_CHARS:
                truncation_warnings.append(f"{file.name} ({len(text):,} characters — only the first {MAX_CHARS:,} were analyzed)")
            truncated = text[:MAX_CHARS]
            prompt = build_prompt(truncated, active_fields)

            progress.progress((idx + 0.5) / len(uploaded_files), text=f"Extracting from {file.name}...")
            try:
                raw = call_openai(prompt, OPENAI_API_KEY, OPENAI_MODEL)
            except Exception as e:
                rows.append({"File": file.name, "Error": f"Extraction failed: {e}"})
                continue

            parsed = parse_json_response(raw)
            rows.extend(build_result_rows(file.name, parsed, active_fields))
            source_items[file.name] = build_source_items(file.name, parsed, active_fields)

        progress.progress(1.0, text="Done")
        st.session_state.results = pd.DataFrame(rows)
        st.session_state.source_items = source_items
        st.session_state.pdf_bytes = pdf_bytes_map
        st.session_state.selected_source = None
        st.session_state.truncation_warnings = truncation_warnings
        st.session_state.reviewable_columns = [f["name"] for f in active_fields]

    if st.session_state.results is not None:
        if st.session_state.get("truncation_warnings"):
            for w in st.session_state.truncation_warnings:
                st.warning(f"⚠️ {w} — terms appearing later in the document may have been missed.")

        # Apply any reviewer edits — every cell (except File) can be corrected,
        # not just ones the model marked "Not found" — so both the table and
        # the Excel export always reflect the reviewer's current word on it.
        display_df = apply_reviewer_overrides(st.session_state.results)
        editable_cols = [c for c in display_df.columns if c not in ("File", "Error")]
        for row_idx in display_df.index:
            for col in editable_cols:
                override_key = f"override_{row_idx}_{col}"
                if override_key in st.session_state:
                    display_df.at[row_idx, col] = st.session_state[override_key]

        st.markdown('<div class="step-label">Results</div>', unsafe_allow_html=True)
        st.markdown(
            display_df.to_html(index=False, escape=False, na_rep=""),
            unsafe_allow_html=True,
        )

        buffer = io.BytesIO()
        display_df.to_excel(buffer, index=False, engine="openpyxl")
        st.download_button(
            "Download as Excel",
            data=buffer.getvalue(),
            file_name="contract_key_terms.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # -----------------------------------------------------------------
        # Edit values — expand a row to correct anything, not just blanks
        # -----------------------------------------------------------------
        st.markdown('<div class="step-label">Edit values</div>', unsafe_allow_html=True)
        st.caption("Open a row to correct any value — whether it's missing or you simply disagree with what was extracted. Changes apply to the table and Excel export immediately.")
        for row_idx in display_df.index:
            file_label = display_df.at[row_idx, "File"] if "File" in display_df.columns else f"Row {row_idx}"
            channel_val = display_df.at[row_idx, "Channel"] if "Channel" in display_df.columns else None
            channel_suffix = f" — {channel_val}" if channel_val and str(channel_val).strip() else ""
            has_missing = any(
                str(st.session_state.results.at[row_idx, c]).strip().lower() == "not found"
                for c in editable_cols if c in st.session_state.results.columns
            )
            with st.expander(f"{file_label}{channel_suffix}" + ("  ⚠️ has missing values" if has_missing else "")):
                for col in editable_cols:
                    if col not in display_df.columns:
                        continue
                    override_key = f"override_{row_idx}_{col}"
                    st.text_input(
                        col,
                        value=str(st.session_state.results.at[row_idx, col]),
                        key=override_key,
                    )

        # -----------------------------------------------------------------
        # Verify a value against its source in the PDF
        # -----------------------------------------------------------------
        any_sources = any(st.session_state.source_items.get(fn) for fn in st.session_state.source_items)
        if any_sources:
            st.markdown('<div class="step-label">Verify a value</div>', unsafe_allow_html=True)
            st.caption("Click any extracted value below to see exactly where it came from in the contract.")

            for file_name, items in st.session_state.source_items.items():
                if not items:
                    continue
                with st.expander(file_name, expanded=len(st.session_state.source_items) == 1):
                    col_list, col_preview = st.columns([2, 3])
                    with col_list:
                        for i, item in enumerate(items):
                            has_quote = bool(item["quote"])
                            st.button(
                                f"{item['label']}: {item['value']}" + ("" if has_quote else "  (no source found)"),
                                key=f"src_{file_name}_{i}",
                                use_container_width=True,
                                disabled=not has_quote,
                                on_click=(lambda fn=file_name, idx=i: st.session_state.__setitem__("selected_source", (fn, idx))) if has_quote else None,
                            )

                    with col_preview:
                        sel = st.session_state.selected_source
                        if sel and sel[0] == file_name:
                            item = items[sel[1]]
                            pdf_bytes = st.session_state.pdf_bytes.get(file_name)
                            located = locate_quote(pdf_bytes, item["quote"]) if pdf_bytes else None
                            if located:
                                page_num, boxes = located
                                png = render_highlighted_page(pdf_bytes, page_num, boxes)
                                st.image(png, caption=f"Page {page_num + 1}", use_container_width=True)
                            else:
                                st.info(f"Couldn't pinpoint this on the page. Quoted text: \u201c{item['quote']}\u201d")
                        else:
                            st.caption("Select a value on the left to preview its source here.")

        st.markdown('<div class="step-label">Confirm</div>', unsafe_allow_html=True)
        st.caption("Locks in the table above — including any manual corrections — as the data every later step will use.")
        if st.button("✅ Confirm & Continue to Classification", type="primary"):
            st.session_state.confirmed_audit_df = apply_reviewer_overrides(st.session_state.results).copy()
            st.session_state.wizard_step = 2
            st.session_state.max_unlocked_step = max(st.session_state.max_unlocked_step, 2)
            st.rerun()

        if st.session_state.get("confirmed_audit_df") is not None:
            st.success("Confirmed ✓ — Classification and Forecasting will use this snapshot until you confirm again.")

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


CLASSIFICATION_CATEGORIES = {
    "HC": "Human Capital",
    "CO": "Incentives for Entrepreneurship",
    "CS": "Corporate Services",
    "GS": "Growth Strategy",
    "MM": "Marketing and Communication",
}


def build_classification_prompt(contract_text: str) -> str:
    category_lines = "\n".join(f'  - "{abbr}" = {full}' for abbr, full in CLASSIFICATION_CATEGORIES.items())
    return f"""You are classifying a contract along two dimensions. Read the contract text below and
return ONLY a JSON object with exactly this shape:

{{
  "pricing_classification": "Fixed" | "Variable" | "Mixed" | "Not found",
  "category": "HC" | "CO" | "CS" | "GS" | "MM",
  "category_reasoning": "<one short phrase, under 12 words, on why this category fits>"
}}

Rules:
- Classify "pricing_classification" as "Variable" if fees depend on usage/volume (e.g. per-click, per-impression, per-lead, commission-based), "Fixed" if it's a flat recurring or one-time fee, "Mixed" if both appear, "Not found" if unclear.
- Classify "category" as exactly one of the following abbreviations, based on the contract's subject matter and the nature of the services/goods being provided:
{category_lines}
- Pick the single best-fitting category even if the contract could plausibly touch more than one — choose based on the PRIMARY subject matter of the agreement.
- Return ONLY the JSON object, no other text, no markdown fences.

CONTRACT TEXT:
\"\"\"
{contract_text}
\"\"\"
"""


def render_classification_mode():
    pdf_bytes_map = st.session_state.get("pdf_bytes", {})

    if not pdf_bytes_map:
        st.info(
            "No contracts loaded yet. Upload and run extraction in **Contract Audit** mode first — "
            "Classification mode automatically reuses those same files."
        )
        return

    st.markdown('<div class="step-label">Contracts loaded from Contract Audit</div>', unsafe_allow_html=True)
    st.caption(", ".join(pdf_bytes_map.keys()))

    if "classification_results" not in st.session_state:
        st.session_state.classification_results = None

    st.markdown('<div class="step-label">Classify</div>', unsafe_allow_html=True)
    if st.button(
        "Classify contracts",
        type="primary",
        disabled=not OPENAI_API_KEY,
    ):
        results = []
        file_names = list(pdf_bytes_map.keys())
        progress = st.progress(0.0, text="Starting...")

        for idx, file_name in enumerate(file_names):
            progress.progress(idx / len(file_names), text=f"Reading {file_name}...")
            try:
                text = extract_pdf_text(io.BytesIO(pdf_bytes_map[file_name]))
            except Exception as e:
                results.append({"file": file_name, "error": f"Failed to read PDF: {e}"})
                continue

            if not text.strip():
                results.append({"file": file_name, "error": "No extractable text (likely a scanned/image PDF)"})
                continue

            prompt = build_classification_prompt(text[:MAX_CHARS])
            progress.progress((idx + 0.5) / len(file_names), text=f"Classifying {file_name}...")
            try:
                raw = call_openai(prompt, OPENAI_API_KEY, OPENAI_MODEL)
            except Exception as e:
                results.append({"file": file_name, "error": f"Classification failed: {e}"})
                continue

            parsed = parse_json_response(raw)
            if "error" in parsed:
                results.append({"file": file_name, "error": parsed["error"]})
                continue

            results.append({
                "file": file_name,
                "pricing_classification": parsed.get("pricing_classification", "Not found"),
                "category": parsed.get("category", "") if parsed.get("category") in CLASSIFICATION_CATEGORIES else "",
                "category_reasoning": parsed.get("category_reasoning", ""),
            })

        progress.progress(1.0, text="Done")
        st.session_state.classification_results = results

    results = st.session_state.classification_results
    if not results:
        return

    st.markdown('<div class="step-label">Results</div>', unsafe_allow_html=True)

    header = st.columns([2.2, 1.5, 1.8, 1.8])
    header[0].markdown("**File**")
    header[1].markdown("**Pricing Classification**")
    header[2].markdown("**Suggested Category**")
    header[3].markdown("**Reviewer Classification**")

    category_options = list(CLASSIFICATION_CATEGORIES.keys())
    export_rows = []

    for r in results:
        row = st.columns([2.2, 1.5, 1.8, 1.8])
        row[0].write(r["file"])

        if "error" in r:
            row[1].write("—")
            row[2].write("—")
            row[3].write(r["error"])
            export_rows.append({
                "File": r["file"], "Pricing Classification": "Error",
                "Suggested Category": "", "Reviewer Classification": "", "Notes": r["error"],
            })
            continue

        row[1].write(r["pricing_classification"])
        suggested = r["category"]
        suggested_label = f"{suggested} — {CLASSIFICATION_CATEGORIES[suggested]}" if suggested else "Not classified"
        row[2].write(suggested_label)

        default_index = category_options.index(suggested) if suggested in category_options else 0
        reviewer_choice = row[3].selectbox(
            "Reviewer override",
            category_options,
            index=default_index,
            key=f"reviewer_class_{r['file']}",
            label_visibility="collapsed",
            format_func=lambda abbr: f"{abbr} — {CLASSIFICATION_CATEGORIES[abbr]}",
        )

        export_rows.append({
            "File": r["file"],
            "Pricing Classification": r["pricing_classification"],
            "Suggested Category": suggested,
            "Reviewer Classification": reviewer_choice,
            "Notes": r.get("category_reasoning", ""),
        })

    st.caption("Reviewer Classification defaults to the suggested category — change it if the model got it wrong.")

    export_df = pd.DataFrame(export_rows)
    buffer = io.BytesIO()
    export_df.to_excel(buffer, index=False, engine="openpyxl")
    st.download_button(
        "Download as Excel",
        data=buffer.getvalue(),
        file_name="contract_classification.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown('<div class="step-label">Confirm</div>', unsafe_allow_html=True)
    st.caption("Locks in each contract's final category above — including any reviewer overrides — for Forecasting to use.")
    if st.button("✅ Confirm & Continue", type="primary"):
        st.session_state.confirmed_classification = [
            {"file": r["File"], "category": r["Reviewer Classification"]}
            for r in export_rows if r.get("Reviewer Classification")
        ]
        st.session_state.wizard_step = 3
        st.session_state.max_unlocked_step = max(st.session_state.max_unlocked_step, 3)
        st.rerun()

    if st.session_state.get("confirmed_classification") is not None:
        st.success("Confirmed ✓ — Forecasting will use these categories until you confirm again.")

def guess_column(columns, keywords):
    for col in columns:
        norm = str(col).strip().lower()
        if any(kw in norm for kw in keywords):
            return col
    return None


def _looks_like_header_label(value) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    text = str(value).strip()
    if not text or text.lower() in ("nan", "none"):
        return False
    lower = text.lower()
    if re.search(r"(date|amount|account|vendor|supplier|name|item|desc|value|cost|spend|debit)", lower):
        return True
    if re.fullmatch(r"[A-Za-z][A-Za-z /_&-]{2,40}", text) and not re.search(r"llc|fz-|fze", lower):
        return True
    return False


def _header_row_score(values) -> int:
    labels = [v for v in values]
    score = 0
    headerish = 0
    for v in labels:
        if _looks_like_header_label(v):
            headerish += 1
            score += 3
        text = "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
        lower = text.lower()
        if any(k in lower for k in ("date", "posted", "amount", "account name", "vendor", "supplier")):
            score += 6
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed) and re.search(r"\d{4}", text):
            score -= 8
        if re.fullmatch(r"-?\d+(\.\d+)?", text.replace(",", "")):
            score -= 4
        if re.search(r"\b(llc|fz-llc|fze|limited)\b", lower):
            score -= 6
    if headerish >= 3:
        score += 8
    return score


def _series_date_hits(series) -> int:
    hits = 0
    for v in series.dropna().head(40):
        parsed = pd.to_datetime(v, errors="coerce", dayfirst=True)
        if pd.isna(parsed):
            parsed = pd.to_datetime(v, errors="coerce")
        if pd.notna(parsed) and 2018 <= getattr(parsed, "year", 0) <= 2040:
            hits += 1
    return hits


def _series_amount_score(series) -> float:
    nums = pd.to_numeric(series, errors="coerce").dropna()
    if len(nums) < 2:
        return -1
    # Account codes often repeat; amounts vary and are not all 4–6 digit IDs.
    unique_ratio = nums.nunique() / max(len(nums), 1)
    mean_abs = float(nums.abs().mean())
    if unique_ratio < 0.25 and mean_abs > 1000:
        return unique_ratio
    return unique_ratio * (1 + min(mean_abs / 10000.0, 5))


def _series_account_score(series) -> float:
    texts = series.dropna().astype(str).str.strip()
    texts = texts[~texts.str.lower().isin(["nan", "none", ""])]
    if texts.empty:
        return -1
    score = 0.0
    llc_hits = texts.str.contains(r"llc|fz|limited|consult", case=False, regex=True).mean()
    score += llc_hits * 5
    score += min(texts.nunique() / max(len(texts), 1), 1) * 2
    return score


def infer_actuals_columns(df):
    """Pick date / amount / account / line-item columns from cell values when
    headers were actually a data row (dates, vendor names, amounts as titles)."""
    date_col = amount_col = account_col = item_col = None
    best_dates, best_amt, best_acct = -1, -1, -1
    for col in df.columns:
        d_hits = _series_date_hits(df[col])
        if d_hits > best_dates:
            best_dates, date_col = d_hits, col
        a_score = _series_amount_score(df[col])
        if a_score > best_amt:
            best_amt, amount_col = a_score, col
        ac_score = _series_account_score(df[col])
        if ac_score > best_acct:
            best_acct, account_col = ac_score, col
    # Don't let the date column also win amount/account.
    for col in df.columns:
        if col in (date_col, amount_col, account_col):
            continue
        if guess_column([col], ["item", "desc", "vendor", "particular", "narration", "details"]):
            item_col = col
            break
    if item_col is None:
        for col in df.columns:
            if col not in (date_col, amount_col, account_col):
                if df[col].dtype == object:
                    item_col = col
                    break
    if best_dates < 2:
        date_col = None
    if best_amt <= 0:
        amount_col = None
    if best_acct < 1:
        account_col = None
    return date_col, amount_col, account_col, item_col


def load_actuals_table(dump_file, sheet_name=None):
    """Read CSV/Excel, find the real header row, fall back to generic names + inference."""
    if hasattr(dump_file, "seek"):
        dump_file.seek(0)
    if dump_file.name.lower().endswith(".csv"):
        raw = pd.read_csv(dump_file, header=None)
        return _frame_from_raw(raw), None

    xls = pd.ExcelFile(dump_file, engine="openpyxl")
    if sheet_name is None:
        sheet_name = xls.sheet_names[0]
    raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)
    return _frame_from_raw(raw), sheet_name


def _frame_from_raw(raw: pd.DataFrame) -> pd.DataFrame:
    if raw is None or raw.empty:
        return pd.DataFrame()
    best_i, best_score = 0, -10**9
    scan = min(12, len(raw))
    for i in range(scan):
        score = _header_row_score(list(raw.iloc[i].values))
        if score > best_score:
            best_score, best_i = score, i
    if best_score >= 8:
        columns = []
        seen = {}
        for i, v in enumerate(raw.iloc[best_i].tolist()):
            label = str(v).strip() if pd.notna(v) else f"Column {i+1}"
            if label.lower() in ("nan", "none", ""):
                label = f"Column {i+1}"
            if label in seen:
                seen[label] += 1
                label = f"{label} ({seen[label]})"
            else:
                seen[label] = 1
            columns.append(label)
        df = raw.iloc[best_i + 1 :].copy()
        df.columns = columns
    else:
        df = raw.copy()
        df.columns = [f"Column {i+1}" for i in range(df.shape[1])]
    df = df.reset_index(drop=True)
    df = df.dropna(how="all")
    return _flatten_columns(df)


_LEGAL_NAME_TOKENS = {
    "llc", "l", "fz", "fzc", "fzco", "fze", "ltd", "limited", "co", "company",
    "pjsc", "psc", "inc", "corp", "corporation", "llp", "plc", "sa", "sarl",
    "gmbh", "the", "and", "of",
}


def normalize_name(x) -> str:
    text = str(x).strip().lower()
    text = text.replace("l.l.c.", "llc").replace("l.l.c", "llc")
    text = re.sub(r"[^\w\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def significant_name_tokens(x) -> set:
    """Drops legal suffixes so 'Pulse Media FZ-LLC' and 'Meridian … FZ-LLC'
    do not look like the same vendor."""
    return {t for t in normalize_name(x).split() if t not in _LEGAL_NAME_TOKENS and len(t) > 1}


def _flatten_columns(df):
    out = df.copy()
    if isinstance(out.columns, pd.MultiIndex):
        out.columns = [
            " ".join(str(part) for part in tup if str(part).strip() and str(part).lower() not in ("nan", "none", "unnamed: 0"))
            for tup in out.columns
        ]
    out.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in out.columns]
    return out


def _month_number_from_label(label):
    if label is None or (isinstance(label, float) and pd.isna(label)):
        return None
    if isinstance(label, (pd.Timestamp,)):
        return int(label.month)
    try:
        import datetime as _dt
        if isinstance(label, _dt.datetime):
            return int(label.month)
    except Exception:
        pass

    raw = str(label).strip()
    raw_l = raw.lower()
    if raw_l.startswith("unnamed") or raw_l in ("nan", "none", ""):
        return None

    # Excel month headers often arrive as "2026-06-01 00:00:00"
    parsed = pd.to_datetime(raw, errors="coerce")
    if pd.notna(parsed) and re.search(r"\d{4}|\d{1,2}\s*[-/]\s*\d", raw):
        return int(parsed.month)

    compact = re.sub(r"[^a-z0-9]", "", raw_l)
    if not compact:
        return None
    full = [
        "january", "february", "march", "april", "may", "june",
        "july", "august", "september", "october", "november", "december",
    ]
    short = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    for i, (name, abbr) in enumerate(zip(full, short), start=1):
        if compact == name or compact == abbr:
            return i
        if compact.startswith(name) or compact.startswith(abbr):
            rest = compact[len(name):] if compact.startswith(name) else compact[len(abbr):]
            if rest == "" or rest.isdigit():
                return i
        if re.search(rf"\b{abbr}\b", raw_l) or re.search(rf"\b{name}\b", raw_l):
            return i
    if compact.isdigit() and 1 <= int(compact) <= 12:
        return int(compact)
    return None


def _add_lookup_value(lookup, name, month_num, val):
    if not name or month_num is None:
        return
    try:
        number = float(val)
    except (TypeError, ValueError):
        return
    if pd.isna(number):
        return
    key = (name, month_num)
    lookup[key] = lookup.get(key, 0.0) + number


def build_monthly_actuals_lookup(df_dump, date_col, name_col, amount_col, year):
    """Builds {(normalized_name, month_num): amount} from a transaction dump
    or from a month-column pivot (Account Name × Jan/Feb/Mar)."""
    df = _flatten_columns(df_dump)
    lookup = {}

    candidate_name_cols = []
    for col in (name_col,):
        if col and col in df.columns:
            candidate_name_cols.append(col)
    for col in df.columns:
        cl = str(col).lower()
        if any(word in cl for word in ("account", "name", "vendor", "supplier", "counterparty", "entity")):
            if col not in candidate_name_cols:
                candidate_name_cols.append(col)

    if date_col in df.columns and amount_col in df.columns:
        dated = df.copy()
        dated["_date"] = pd.to_datetime(dated[date_col], errors="coerce", dayfirst=True)
        if dated["_date"].isna().mean() > 0.5:
            dated["_date"] = pd.to_datetime(dated[date_col], errors="coerce", dayfirst=False)
        usable = dated[dated["_date"].notna()]
        if year:
            year_filtered = usable[usable["_date"].dt.year == year]
            if not year_filtered.empty:
                usable = year_filtered
        for ncol in candidate_name_cols:
            if ncol not in usable.columns:
                continue
            tmp = usable.copy()
            tmp["_name_norm"] = tmp[ncol].apply(normalize_name)
            tmp["_month"] = tmp["_date"].dt.month
            tmp[amount_col] = pd.to_numeric(tmp[amount_col], errors="coerce")
            tmp = tmp.dropna(subset=["_name_norm", amount_col])
            for (nm, m), val in tmp.groupby(["_name_norm", "_month"])[amount_col].sum().items():
                _add_lookup_value(lookup, nm, m, val)

    month_cols = {}
    for col in df.columns:
        month_num = _month_number_from_label(col)
        if month_num:
            month_cols[col] = month_num

    if month_cols:
        name_cols = candidate_name_cols or [c for c in df.columns if c not in month_cols]
        for ncol in name_cols:
            if ncol not in df.columns or ncol in month_cols:
                continue
            for _, row in df.iterrows():
                name = normalize_name(row.get(ncol, ""))
                if not name or name in ("nan", "none", "total", "grand total"):
                    continue
                for col, month_num in month_cols.items():
                    _add_lookup_value(lookup, name, month_num, row[col])

    return lookup


def lookup_actual_amount(actuals_lookup, name_candidates, month):
    if not actuals_lookup:
        return None
    norms = []
    cores = []
    for n in name_candidates:
        if n is None or str(n).strip() == "":
            continue
        norms.append(normalize_name(n))
        cores.append(significant_name_tokens(n))
    for n in norms:
        if (n, month) in actuals_lookup:
            return actuals_lookup[(n, month)]
    # Distinctive-token match only. "fz" + "llc" must never pair Pulse with Meridian.
    for (nm, m), val in actuals_lookup.items():
        if m != month:
            continue
        nm_core = significant_name_tokens(nm)
        if not nm_core:
            continue
        for core in cores:
            if not core:
                continue
            if core == nm_core:
                return val
            if core <= nm_core or nm_core <= core:
                return val
            overlap = core & nm_core
            need = 2 if min(len(core), len(nm_core)) >= 2 else 1
            if len(overlap) >= need and len(overlap) >= min(len(core), len(nm_core)) - 0:
                # Require at least half of the shorter distinctive name.
                if len(overlap) * 2 >= min(len(core), len(nm_core)):
                    return val
    return None


CATEGORY_ABBREVIATIONS = list(CLASSIFICATION_CATEGORIES.keys())


def classify_row_color(cell) -> str:
    """Returns 'purple' (heading), 'grey' (subheading), or 'plain' (line item),
    based on the theme-based fill this budget template uses."""
    fg = cell.fill.fgColor
    if fg.type == "theme" and fg.theme == 8:
        return "purple"
    elif fg.type == "theme" and fg.theme == 0:
        return "grey"
    return "plain"


def build_sheet_structure(ws) -> list:
    """Walks column A from row 3 onward (skipping the title and month-header rows)
    and classifies each row as heading / subheading / line_item / total."""
    nodes = []
    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        name = cell.value
        if name is None or (isinstance(name, str) and not name.strip()):
            continue
        color = classify_row_color(cell)
        if isinstance(name, str) and name.strip().lower() == "total":
            node_type = "total"
        elif color == "purple":
            node_type = "heading"
        elif color == "grey":
            node_type = "subheading"
        else:
            node_type = "line_item"
        nodes.append({"type": node_type, "row": row, "name": name.strip() if isinstance(name, str) else name})
    return nodes


def get_hierarchy(nodes: list):
    """Builds the heading -> subheading -> line_item tree, plus a flat
    {normalized_line_item_name: (heading, subheading)} index for existence checks."""
    headings = []
    current_heading = None
    current_subheading = None
    line_item_index = {}
    total_row = None
    for node in nodes:
        if node["type"] == "heading":
            current_heading = {"name": node["name"], "row": node["row"], "subheadings": [], "direct_items": []}
            headings.append(current_heading)
            current_subheading = None
        elif node["type"] == "subheading":
            current_subheading = {"name": node["name"], "row": node["row"], "items": []}
            if current_heading is not None:
                current_heading["subheadings"].append(current_subheading)
        elif node["type"] == "line_item":
            key = normalize_name(node["name"])
            heading_name = current_heading["name"] if current_heading else None
            subheading_name = current_subheading["name"] if current_subheading else None
            line_item_index[key] = (heading_name, subheading_name, node["row"])
            if current_subheading is not None:
                current_subheading["items"].append(node)
            elif current_heading is not None:
                current_heading["direct_items"].append(node)
        elif node["type"] == "total":
            total_row = node["row"]
    return headings, line_item_index, total_row


def parse_date_flexible(text):
    if not text or str(text).strip().lower() in ("not found", ""):
        return None
    try:
        result = pd.to_datetime(text, errors="coerce")
        return None if pd.isna(result) else result
    except Exception:
        return None


_WORD_NUMBERS = {
    "a": 1, "an": 1, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11,
    "twelve": 12, "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16,
    "eighteen": 18, "twenty": 20, "twentyfour": 24, "twenty-four": 24,
    "thirty": 30, "thirtysix": 36, "thirty-six": 36,
}


def _number_from_token(token: str):
    token = (token or "").strip().lower()
    if not token:
        return None
    if token.isdigit():
        return int(token)
    if token in _WORD_NUMBERS:
        return _WORD_NUMBERS[token]
    return None


def parse_duration_months(term_text, effective_date=None):
    if not term_text or str(term_text).strip().lower() == "not found":
        return None
    text = str(term_text).lower()

    # Digits or words: "6 months", "six months", "two (2) years", "12-month term".
    patterns = [
        (r"(one|two|three|four|five|six|seven|eight|nine|ten|eleven|twelve|"
         r"thirteen|fourteen|fifteen|sixteen|eighteen|twenty|thirty|"
         r"twenty-four|thirty-six|a|an|\d+)\s*(?:\([^)]{0,12}\)\s*)?"
         r"[- ]*(year|yr)s?\b", 12),
        (r"(one|two|three|four|five|six|seven|eight|nine|ten|eleven|twelve|"
         r"thirteen|fourteen|fifteen|sixteen|eighteen|twenty|thirty|"
         r"twenty-four|thirty-six|a|an|\d+)\s*(?:\([^)]{0,12}\)\s*)?"
         r"[- ]*(month|mo)s?\b", 1),
    ]
    for pattern, multiplier in patterns:
        m = re.search(pattern, text)
        if m:
            n = _number_from_token(m.group(1))
            if n:
                return n * multiplier

    end_date = parse_date_flexible(term_text)
    if end_date is not None and effective_date is not None:
        months = (end_date.year - effective_date.year) * 12 + (end_date.month - effective_date.month) + 1
        return max(months, 1)
    return None


def parse_amount(value_text, duration_months=None):
    """Returns the figure that is spread over the contract term.

    - "AED 42,000 per month" × 6 months → 252,000 (rate × term).
    - "AED 480,000 per annum" on a 24-month term → 480,000 (the stated
      figure is spread over the full term: 480,000 / 24 = 20,000 a month).
      It is NOT doubled into a 960,000 two-year total.
    """
    if not value_text or str(value_text).strip().lower() == "not found":
        return None
    text = str(value_text).replace(",", "")
    m = re.search(r"(\d+\.?\d*)", text)
    if not m:
        return None
    amount = float(m.group(1))

    lower = text.lower()
    if duration_months:
        if re.search(r"per\s*month|/\s*month|monthly", lower):
            return amount * duration_months
        if re.search(r"per\s*quarter|quarterly", lower):
            return amount * (duration_months / 3.0)
        if re.search(r"per\s*week|weekly", lower):
            return amount * (duration_months * 4.345)
    return amount


def first_validity_month(effective_date, calendar_year, actuals_by_month):
    """First 2026 month that counts toward Column B.

    Starts in the effective-date month (March for Pulse). The month before
    that is pulled in only when an actual was posted there (February actual
    on a March start). Otherwise the envelope starts in the start month.
    """
    if effective_date is None:
        return 1
    if effective_date.year > calendar_year:
        return 13
    if effective_date.year < calendar_year:
        return 1
    start_m = int(effective_date.month)
    prior = start_m - 1
    if prior >= 1 and actuals_by_month and prior in actuals_by_month:
        return prior
    return start_m


def contract_end_date(effective_date, duration_months):
    """Inclusive end date: start 21 Jan + 6 months → 20 Jul."""
    if effective_date is None or not duration_months:
        return None
    return pd.Timestamp(effective_date) + pd.DateOffset(months=int(duration_months)) - pd.Timedelta(days=1)


def expiry_month_cap(effective_date, duration_months, calendar_year):
    """Month number used as the forecast-period cap: min(12, expiry month)
    when the contract ends in this year, otherwise 12 if it runs past December."""
    end = contract_end_date(effective_date, duration_months)
    if end is None:
        return 12
    if end.year > calendar_year:
        return 12
    if end.year < calendar_year:
        return 0
    return int(end.month)


def payment_periods_in_year(effective_date, duration_months, calendar_year):
    """How many contractual payment months begin in this calendar year."""
    if effective_date is None or not duration_months:
        return 0
    start = pd.Timestamp(effective_date)
    n = 0
    for i in range(int(duration_months)):
        if (start + pd.DateOffset(months=i)).year == calendar_year:
            n += 1
    return n


def calendar_months_spanned(effective_date, duration_months, calendar_year):
    """Every calendar month the contract touches in this year, including a
    tail month (Jan 21–Jul 20 spans January through July)."""
    end = contract_end_date(effective_date, duration_months)
    if effective_date is None or end is None:
        return []
    months = []
    cursor = pd.Timestamp(year=effective_date.year, month=effective_date.month, day=1)
    last = pd.Timestamp(year=end.year, month=end.month, day=1)
    while cursor <= last:
        if cursor.year == calendar_year:
            months.append(int(cursor.month))
        cursor += pd.DateOffset(months=1)
    return months


def compute_prorated_budget(contract_value, effective_date, duration_months, calendar_year):
    """Returns (total_for_year, monthly_rate, [calendar months spanned]) or (None, None, []).

    Column B uses contractual payment months in the year (6 × monthly), not the
    extra tail calendar month a mid-month end may spill into.
    """
    if contract_value is None or effective_date is None or not duration_months:
        return None, None, []
    monthly_rate = contract_value / duration_months
    periods = payment_periods_in_year(effective_date, duration_months, calendar_year)
    months_spanned = calendar_months_spanned(effective_date, duration_months, calendar_year)
    total = monthly_rate * periods
    return total, monthly_rate, months_spanned


def build_placement_prompt(contract_text: str, headings: list) -> str:
    heading_lines = []
    for h in headings:
        if h["subheadings"]:
            subs = ", ".join(f'"{s["name"]}"' for s in h["subheadings"])
            heading_lines.append(f'  - "{h["name"]}" (subheadings: {subs})')
        else:
            heading_lines.append(f'  - "{h["name"]}" (no subheadings — items sit directly under it)')
    heading_block = "\n".join(heading_lines)
    return f"""A new contract needs to be placed into one of the existing budget categories below.
Read the contract text and pick the single best-fitting heading (and subheading, if that
heading has any) based on what the contract is actually for.

Existing categories:
{heading_block}

Return ONLY a JSON object of this shape:
{{
  "heading": "<one of the exact heading names above>",
  "subheading": "<one of that heading's exact subheading names, or empty string if the heading has no subheadings>"
}}

CONTRACT TEXT:
\"\"\"
{contract_text}
\"\"\"
"""


def find_insertion_row(heading: dict, subheading_name: str) -> int:
    """Row just after the last existing item in the target group (or right after
    the heading/subheading itself if that group has no items yet)."""
    if subheading_name:
        for sub in heading["subheadings"]:
            if sub["name"] == subheading_name:
                return (sub["items"][-1]["row"] + 1) if sub["items"] else (sub["row"] + 1)
        return heading["row"] + 1
    if heading["direct_items"]:
        return heading["direct_items"][-1]["row"] + 1
    if heading["subheadings"]:
        last_sub = heading["subheadings"][-1]
        return (last_sub["items"][-1]["row"] + 1) if last_sub["items"] else (last_sub["row"] + 1)
    return heading["row"] + 1


def write_forecast_workbook(wb, plan_by_sheet: dict) -> bytes:
    """Inserts new line items per sheet (bottom-to-top so row numbers stay valid),
    copies formatting from a neighboring line-item row, writes the prorated Budget
    figure and its monthly spread, then rewrites SUM formulas for every
    heading/subheading/grand-total row so the workbook stays a live, editable file."""
    from openpyxl.utils import get_column_letter
    import copy as _copy

    MONTH_COL_START = 4  # column D = Jan

    for sheet_name, insertions in plan_by_sheet.items():
        if not insertions:
            continue
        ws = wb[sheet_name]

        def _write_line(ws, row, ins):
            ws.cell(row=row, column=1).value = ins["counterparty"]
            ws.cell(row=row, column=2).value = round(ins["prorated_total"], 2) if ins["prorated_total"] else 0
            ins_monthly_values = ins.get("monthly_values") or {}
            for m in range(1, 13):
                col = MONTH_COL_START + (m - 1)
                # Only the computed actual/forecast map — never back-fill a
                # flat monthly rate into Jan–selected-month cells.
                ws.cell(row=row, column=col).value = max(round(float(ins_monthly_values.get(m, 0) or 0), 2), 0)

        updates = [ins for ins in insertions if ins.get("existing_row")]
        creates = [ins for ins in insertions if not ins.get("existing_row")]
        for ins in updates:
            _write_line(ws, ins["existing_row"], ins)
        for ins in sorted(creates, key=lambda x: x["insertion_row"], reverse=True):
            row = ins["insertion_row"]
            ws.insert_rows(row, 1)
            template_row = row - 1 if row > 1 else row + 1
            for col in range(1, 17):
                src = ws.cell(row=template_row, column=col)
                dst = ws.cell(row=row, column=col)
                dst.font = _copy.copy(src.font)
                dst.fill = _copy.copy(src.fill)
                dst.border = _copy.copy(src.border)
                dst.number_format = src.number_format
                dst.alignment = _copy.copy(src.alignment)
            _write_line(ws, row, ins)

        # Recompute the hierarchy now that rows have shifted, and rewrite rollup formulas.
        nodes = build_sheet_structure(ws)
        headings, _, total_row = get_hierarchy(nodes)

        heading_cell_refs = []
        for h in headings:
            if h["subheadings"]:
                sub_refs = []
                for sub in h["subheadings"]:
                    if sub["items"]:
                        first, last = sub["items"][0]["row"], sub["items"][-1]["row"]
                        for col in [2] + list(range(4, 16)):
                            letter = get_column_letter(col)
                            ws.cell(row=sub["row"], column=col).value = f"=SUM({letter}{first}:{letter}{last})"
                    sub_refs.append(sub["row"])
                for col in [2] + list(range(4, 16)):
                    letter = get_column_letter(col)
                    formula = "=" + "+".join(f"{letter}{r}" for r in sub_refs) if sub_refs else 0
                    ws.cell(row=h["row"], column=col).value = formula
            elif h["direct_items"]:
                first, last = h["direct_items"][0]["row"], h["direct_items"][-1]["row"]
                for col in [2] + list(range(4, 16)):
                    letter = get_column_letter(col)
                    ws.cell(row=h["row"], column=col).value = f"=SUM({letter}{first}:{letter}{last})"
            heading_cell_refs.append(h["row"])

        if total_row is None:
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=1).value = "Total"
        for col in [2] + list(range(4, 16)):
            letter = get_column_letter(col)
            formula = "=" + "+".join(f"{letter}{r}" for r in heading_cell_refs) if heading_cell_refs else 0
            ws.cell(row=total_row, column=col).value = formula

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_actuals_upload_mode():
    st.caption("Upload a transaction-level actuals export here — Forecasting mode will automatically use it instead of the budget file's own month columns. This step is optional — skip straight to Forecasting if you don't need it.")

    existing = st.session_state.get("actuals_dump_mapping")
    if existing:
        df_dump, date_col, dump_name_col, amount_col, account_col = existing
        account_note = f", account name from '{account_col}'" if account_col else ""
        st.success(f"Actuals dump loaded — {len(df_dump):,} rows, matched on '{dump_name_col}' with amounts from '{amount_col}'{account_note}.")
        if st.button("Clear uploaded actuals dump"):
            st.session_state.pop("actuals_dump_mapping", None)
            st.rerun()
        st.markdown('<div class="step-label">Replace it</div>', unsafe_allow_html=True)

    dump_file = st.file_uploader("Actuals dump (.xlsx or .csv)", type=["xlsx", "csv"], key="actuals_dump_file")

    if not dump_file:
        if not existing:
            st.info("Upload a file with one row per transaction, with a date, a line-item name, and an amount — or just continue if you don't have one.")
    else:
        try:
            sheet_name = None
            if not dump_file.name.lower().endswith(".csv"):
                if hasattr(dump_file, "seek"):
                    dump_file.seek(0)
                xls = pd.ExcelFile(dump_file, engine="openpyxl")
                if len(xls.sheet_names) > 1:
                    st.warning(
                        f"This workbook has {len(xls.sheet_names)} sheets — pick the one with **raw, one-row-per-transaction data**."
                    )
                    sheet_name = st.selectbox("Sheet", xls.sheet_names, key="actuals_dump_sheet")
                else:
                    sheet_name = xls.sheet_names[0]
            if hasattr(dump_file, "seek"):
                dump_file.seek(0)
            df_dump, sheet_name = load_actuals_table(dump_file, sheet_name)
        except Exception as e:
            st.error(f"Couldn't read this file: {e}")
            df_dump = None

        if df_dump is not None and df_dump.empty:
            st.warning("This file appears to be empty.")
        elif df_dump is not None:
            st.caption("Preview — headers are detected automatically. If a data row is still showing as a title, set the four boxes from the values you can see.")
            st.markdown(df_dump.head(8).to_html(index=False), unsafe_allow_html=True)
            st.markdown('<div class="step-label">Confirm columns</div>', unsafe_allow_html=True)
            dump_columns = list(df_dump.columns)
            inferred_date, inferred_amount, inferred_account, inferred_item = infer_actuals_columns(df_dump)
            guessed_date = guess_column(dump_columns, ["date", "posted", "transaction"]) or inferred_date
            guessed_dump_name = guess_column(dump_columns, ["name", "item", "channel", "department", "category", "line", "vendor", "desc"]) or inferred_item
            guessed_amount = guess_column(dump_columns, ["amount", "value", "cost", "spend", "debit"]) or inferred_amount
            guessed_account = guess_column(dump_columns, ["account name", "account", "vendor", "supplier", "counterparty"]) or inferred_account

            dc1, dc2, dc3, dc4 = st.columns(4)
            date_col = dc1.selectbox("Date column", dump_columns, index=dump_columns.index(guessed_date) if guessed_date in dump_columns else 0)
            dump_name_col = dc2.selectbox("Line item column", dump_columns, index=dump_columns.index(guessed_dump_name) if guessed_dump_name in dump_columns else 0)
            amount_col = dc3.selectbox("Amount column", dump_columns, index=dump_columns.index(guessed_amount) if guessed_amount in dump_columns else 0)
            account_options = ["(none)"] + dump_columns
            account_default = guessed_account if guessed_account in dump_columns else "(none)"
            account_picked = dc4.selectbox("Account Name column", account_options, index=account_options.index(account_default) if account_default in account_options else 0)
            account_col = None if account_picked == "(none)" else account_picked
            st.caption(
                f"Detected → date: **{date_col}**, amount: **{amount_col}**, "
                f"account: **{account_col or '(none)'}**. "
                f"Account Name must be the column that contains **PULSE MEDIA FZ-LLC**."
            )

            st.session_state["actuals_dump_mapping"] = (df_dump, date_col, dump_name_col, amount_col, account_col)
            st.caption("Matched against your budget file's line items by name in Forecasting mode.")

    st.markdown('<div class="step-label">Continue</div>', unsafe_allow_html=True)
    if st.button("Continue to Forecasting →", type="primary"):
        mapping = st.session_state.get("actuals_dump_mapping")
        if mapping:
            df_dump, date_col, dump_name_col, amount_col, account_col = mapping
            if account_col:
                names = df_dump[account_col].dropna().astype(str).str.strip()
                st.session_state.confirmed_account_names = sorted(set(n for n in names if n))
            else:
                st.session_state.confirmed_account_names = []
        else:
            st.session_state.confirmed_account_names = []
        st.session_state.wizard_step = 4
        st.session_state.max_unlocked_step = max(st.session_state.max_unlocked_step, 4)
        st.rerun()


def first_ready_month(effective_date, calendar_year):
    """First calendar month the contract enters the 2026 (or chosen-year) envelope.

    A mid-month start (e.g. 20 April) is NOT in the envelope yet — it is first
    budgeted from the following month — unless an actual is later found in the
    start month, which pulls that month in. Selecting April (= as-of 30 April)
    with a 20 April start and no April actual means the contract is not live.
    """
    if effective_date is None:
        return None
    if effective_date.day <= 1:
        ready = pd.Timestamp(year=effective_date.year, month=effective_date.month, day=1)
    else:
        ready = pd.Timestamp(year=effective_date.year, month=effective_date.month, day=1) + pd.DateOffset(months=1)
    if ready.year > calendar_year:
        return 13
    if ready.year < calendar_year:
        return 1
    return int(ready.month)


def compute_reforecast(contract_value, effective_date, duration_months, calendar_year, actuals_lookup, name_candidates, through_month_idx, budget_method="fixed"):
    """Fixed budgeting (selected month = close of that month, e.g. May → 31 May):

    Actual period   = January .. selected month (inclusive) — posted actuals only.
    Forecast period = month after selected .. min(December, contract expiry month).

    Column B = monthly rate × contractual payment months that begin in the year
    (Jan 21–Jul 20, 6 months × 42,000 = 252,000), even if January has no actual.

    Forecast monthly rate =
        (Column B − cumulative actuals through selected month)
        / (min(12, expiry_month) − selected_month).

    Months after expiry are always 0. No negatives.

    Flexible: leftover full contract value spread over remaining term months.
    """
    empty = {
        "live": False, "cumulative_actual": 0.0, "actual_months": 0,
        "remaining_period": 0, "orig_2026_budget": 0.0, "orig_monthly_rate": 0.0,
        "method": budget_method,
    }

    year_total, flat_rate, spanned_months = compute_prorated_budget(
        contract_value, effective_date, duration_months, calendar_year
    )
    if not spanned_months or flat_rate is None or contract_value is None or not duration_months:
        return None, None, [], {}, empty

    actuals_by_month = {}
    if actuals_lookup:
        for m in range(1, 13):
            val = lookup_actual_amount(actuals_lookup, name_candidates, m)
            if val is not None:
                actuals_by_month[m] = float(val)

    ready_from = first_validity_month(effective_date, calendar_year, actuals_by_month)
    expiry_cap = expiry_month_cap(effective_date, duration_months, calendar_year)

    # Payment periods that begin in this year (Mar–Dec = 10 for Pulse;
    # Jan–Jun = 6 for Meridian). A prior-month actual adds that month
    # (Feb actual on a March start → 11). The end-date tail month is
    # not an extra payment period.
    periods = payment_periods_in_year(effective_date, duration_months, calendar_year)
    start_m = (
        int(effective_date.month)
        if effective_date is not None and effective_date.year == calendar_year
        else 1
    )
    if ready_from < start_m:
        periods += start_m - ready_from
    year_months = [m for m in range(ready_from, min(12, expiry_cap) + 1)]
    orig_2026_budget = flat_rate * max(periods, 0)
    validity_start_label = MONTH_NAMES[ready_from - 1] if 1 <= ready_from <= 12 else "?"
    validity_end_label = MONTH_NAMES[min(12, expiry_cap) - 1] if 1 <= expiry_cap <= 12 else "Dec"

    monthly_values = {m: 0.0 for m in range(1, 13)}

    # Close of selected month: contract is not in the book yet if its first
    # budget month is still in the future and no start-month actual pulled it in.
    if through_month_idx < ready_from:
        start_label = (
            pd.Timestamp(effective_date).strftime("%d %b %Y")
            if effective_date is not None else "unknown start"
        )
        meta = dict(empty)
        meta["orig_2026_budget"] = orig_2026_budget
        meta["orig_monthly_rate"] = flat_rate
        meta["ready_from"] = ready_from
        meta["reason"] = (
            f"Not in existence at the end of {MONTH_NAMES[through_month_idx - 1]} "
            f"— starts {start_label} and has no actual in the start month. "
            f"First budget month is {MONTH_NAMES[ready_from - 1]}."
        )
        return 0.0, 0.0, [], monthly_values, meta

    actual_period = list(range(1, through_month_idx + 1))
    # Forecast runs only to the contract's expiry month (or December if later).
    forecast_end = max(min(12, expiry_cap), through_month_idx)
    forecast_period = list(range(through_month_idx + 1, forecast_end + 1)) if expiry_cap > through_month_idx else []

    # Cumulative spend is everything posted in the actual period, even if a
    # posting falls outside the envelope months (so overspend is visible).
    cumulative_actual = sum(actuals_by_month.get(m, 0.0) for m in actual_period)
    actual_month_count = sum(1 for m in actual_period if m in actuals_by_month)
    future_months = [m for m in year_months if m in forecast_period]

    insufficient = False
    leftover = 0.0
    if budget_method == "flexible":
        leftover = contract_value - cumulative_actual
        remaining_period = max(int(duration_months) - actual_month_count, 1)
        if leftover < 0:
            insufficient = True
            new_monthly_rate = 0.0
        else:
            new_monthly_rate = leftover / remaining_period
        for m in year_months:
            if m <= through_month_idx:
                monthly_values[m] = max(actuals_by_month.get(m, 0.0), 0.0)
            else:
                monthly_values[m] = new_monthly_rate
        display_total = sum(monthly_values.values())
    else:
        leftover = orig_2026_budget - cumulative_actual
        if leftover < 0:
            insufficient = True

        # (min(12, expiry month) − selected month). Example: expiry July,
        # select May → min(12, 7) − 5 = 2 (June and July).
        remaining_period = max(min(12, expiry_cap) - through_month_idx, 0)
        if remaining_period <= 0:
            new_monthly_rate = 0.0
        elif leftover <= 0:
            new_monthly_rate = 0.0
        else:
            new_monthly_rate = leftover / remaining_period

        if not actuals_lookup:
            for m in year_months:
                monthly_values[m] = flat_rate
            remaining_period = len(year_months)
            new_monthly_rate = flat_rate
        else:
            for m in range(1, 13):
                if m <= through_month_idx:
                    monthly_values[m] = max(actuals_by_month.get(m, 0.0), 0.0)
                elif m in forecast_period and m <= expiry_cap:
                    monthly_values[m] = new_monthly_rate
                else:
                    monthly_values[m] = 0.0

        display_total = orig_2026_budget

    meta = {
        "live": True,
        "cumulative_actual": cumulative_actual,
        "actual_months": actual_month_count,
        "remaining_period": remaining_period,
        "ready_from": ready_from,
        "orig_2026_budget": orig_2026_budget,
        "orig_monthly_rate": flat_rate,
        "method": budget_method,
        "insufficient": insufficient,
        "leftover": leftover,
        "reason": "",
        "actual_period_end": through_month_idx,
        "forecast_months": remaining_period,
        "expiry_month": expiry_cap,
        "periods": periods,
        "duration_months": duration_months,
        "contract_value": contract_value,
        "validity_start": validity_start_label,
        "validity_end": validity_end_label,
    }
    return display_total, new_monthly_rate, year_months, monthly_values, meta


def render_forecast_mode():
    st.caption("Adds new contracts into your Hub71 budget template, pro-rated for the months remaining in the calendar year, and keeps every heading/subheading/grand-total roll-up in sync.")

    if st.session_state.get("confirmed_audit_df") is None:
        st.info("Complete and **Confirm** Contract Audit first — Forecasting reads the confirmed snapshot, not live edits.")
        return
    if not st.session_state.get("confirmed_classification"):
        st.info("Complete and **Confirm** Contract Classification first — Forecasting reads the confirmed snapshot, not live edits.")
        return

    confirmed_names = [r.get("file", "?") for r in st.session_state.confirmed_classification]
    with st.expander(f"Currently confirmed: {len(confirmed_names)} contract(s) — check this is your real, current set", expanded=False):
        for n in confirmed_names:
            st.caption(f"• {n}")
        st.caption("If this shows old test files or contracts you no longer want, go back to Step 1, re-run extraction (it clears the old snapshot), then re-confirm Steps 1 and 2 with the correct files before returning here.")

    st.markdown('<div class="step-label">1 · Upload your budget workbook</div>', unsafe_allow_html=True)
    budget_file = st.file_uploader("Excel file (.xlsx)", type=["xlsx"], label_visibility="collapsed", key="forecast_budget_file")
    if budget_file is not None:
        st.session_state.forecast_budget_bytes = budget_file.getvalue()
        st.session_state.forecast_budget_name = budget_file.name

    budget_bytes = st.session_state.get("forecast_budget_bytes")
    if budget_file is None and budget_bytes:
        st.caption(f"Using last uploaded workbook: **{st.session_state.get('forecast_budget_name', 'budget.xlsx')}** (kept until you Clear or upload a new file).")
    if not budget_bytes:
        st.info("Upload the Hub71 budget workbook (sheets named CO / GS / CS / MM, with purple heading rows and grey subheading rows).")
        if st.session_state.get("forecast_workbook_bytes"):
            st.markdown('<div class="step-label">Last applied forecast</div>', unsafe_allow_html=True)
            result_wb = openpyxl.load_workbook(io.BytesIO(st.session_state.forecast_workbook_bytes), data_only=False)
            for sheet_name in st.session_state.get("forecast_sheets_touched") or result_wb.sheetnames:
                if sheet_name not in result_wb.sheetnames:
                    continue
                ws = result_wb[sheet_name]
                st.caption(f"**{sheet_name}**")
                rows_out = []
                for row in range(3, ws.max_row + 1):
                    name = ws.cell(row=row, column=1).value
                    if name is None:
                        continue
                    budget = ws.cell(row=row, column=2).value
                    month_vals = [ws.cell(row=row, column=4 + m).value for m in range(12)]
                    rows_out.append([name, budget] + month_vals)
                preview_df = pd.DataFrame(rows_out, columns=["Line Item", "Budget"] + MONTH_NAMES)
                st.markdown(preview_df.to_html(index=False, escape=False, na_rep=""), unsafe_allow_html=True)
            st.download_button(
                "Download last updated workbook",
                data=st.session_state.forecast_workbook_bytes,
                file_name="budget_updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        return

    try:
        wb = openpyxl.load_workbook(io.BytesIO(budget_bytes), data_only=False)
    except Exception as e:
        st.error(f"Couldn't read this workbook: {e}")
        return

    # A different uploaded file, actuals dump, calendar year, or "through
    # month" all invalidate any previously generated output — otherwise a
    # stale download could silently persist even after everything upstream
    # has changed. The signature is finalized once through_month is known.
    year = st.number_input("Calendar year this budget covers", min_value=2020, max_value=2100, value=2026, step=1, key="forecast_year")

    import hashlib
    dump_mapping = st.session_state.get("actuals_dump_mapping")
    dump_fingerprint = ""
    if dump_mapping:
        try:
            dump_fingerprint = hashlib.md5(pd.util.hash_pandas_object(dump_mapping[0]).values.tobytes()).hexdigest()
        except Exception:
            dump_fingerprint = str(len(dump_mapping[0]))

    actuals_lookup = None
    through_month_idx = 12
    through_month = None
    budget_method = st.session_state.get("budget_method", "fixed")
    if dump_mapping:
        df_dump, date_col, dump_name_col, amount_col, account_col = dump_mapping
        match_col = account_col or dump_name_col
        actuals_lookup = build_monthly_actuals_lookup(df_dump, date_col, match_col, amount_col, year)
        if account_col and account_col != dump_name_col:
            extra = build_monthly_actuals_lookup(df_dump, date_col, dump_name_col, amount_col, year)
            for k, v in extra.items():
                actuals_lookup.setdefault(k, v)
        budget_method = st.radio(
            "Budgeting method",
            ["fixed", "flexible"],
            index=0,
            horizontal=True,
            key="budget_method",
            format_func=lambda x: "Fixed budgeting (2026 envelope)" if x == "fixed" else "Flexible budgeting (full contract term)",
            help="Fixed: Column B is the locked year envelope. Actual period (Jan–selected month) takes posted actuals only; leftover envelope is spread across the forecast period (next month–Dec) as (B − YTD actuals) / (12 − selected month). Flexible: leftover full contract value is spread over months left on the contract.",
        )
        mc1, mc2 = st.columns([3, 1])
        through_month = mc1.selectbox(
            "Close of month (actuals through)",
            MONTH_NAMES,
            index=0,
            key="forecast_through_month",
            help="Treats the choice as month-end (April = 30 April). Actual period = January through this month — cells come from the actuals file only. Forecast period = the following month through December.",
        )
        through_month_idx = MONTH_NAMES.index(through_month) + 1
        mc2.write("")  # vertical spacer to align button with selectbox
        recomputed = mc2.button("🔄 Recompute", use_container_width=True)
        if recomputed:
            st.success(f"Recomputed using actuals through {through_month} ({year}).")

        if not actuals_lookup:
            st.error(
                "Actuals file is loaded, but no month amounts were read. "
                "In Step 3 pick the sheet that looks like Account Name | Jan | Feb | Mar, "
                "and set Account Name column to 'Account Name'. Then come back and click Recompute."
            )
        with st.expander(f"Debug: actuals lookup ({len(actuals_lookup or {})} entries)", expanded=not actuals_lookup):
            st.caption(f"Matched on column: **{match_col}**")
            if actuals_lookup:
                debug_rows = [{"Name": k[0], "Month": MONTH_NAMES[k[1]-1], "Amount": v} for k, v in sorted(actuals_lookup.items())]
                st.markdown(pd.DataFrame(debug_rows).to_html(index=False), unsafe_allow_html=True)
            else:
                st.warning("Lookup is empty — March cannot pick up 107,208 until this table has rows.")
    else:
        st.caption("No actuals dump loaded (Step 3) — using the flat pro-rated calculation for every contract.")

    file_signature = hashlib.md5(budget_bytes).hexdigest() + f"::{year}::{dump_fingerprint}::{through_month_idx}"
    st.session_state["forecast_budget_signature"] = file_signature

    # Build the sheet structure once so we can match categories to real sheets and show existing headings.
    sheet_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        nodes = build_sheet_structure(ws)
        headings, line_item_index, total_row = get_hierarchy(nodes)
        sheet_data[sheet_name] = {"headings": headings, "line_item_index": line_item_index, "total_row": total_row}

    audit_df = st.session_state.confirmed_audit_df
    pdf_bytes_map = st.session_state.get("pdf_bytes", {})

    st.markdown('<div class="step-label">2 · Review planned changes</div>', unsafe_allow_html=True)

    plan = []  # list of dicts, one per contract needing a decision
    for r in st.session_state.confirmed_classification:
        file_name = r.get("file")
        suggested_category = r.get("category", "")
        category = suggested_category  # already the final reviewer-confirmed category
        if category not in sheet_data or sheet_data.get(category) is None:
            continue  # e.g. HC has no matching sheet in this workbook
        if category not in wb.sheetnames:
            plan.append({"file": file_name, "category": category, "status": "no_sheet"})
            continue

        file_rows = audit_df[audit_df.get("File") == file_name] if "File" in audit_df.columns else pd.DataFrame()
        if file_rows.empty:
            plan.append({"file": file_name, "category": category, "status": "no_audit_data"})
            continue
        row = file_rows.iloc[0]
        counterparty = str(row.get("Counterparty", "")).strip()
        if not counterparty or counterparty.lower() == "not found":
            plan.append({"file": file_name, "category": category, "status": "no_counterparty"})
            continue

        sheet_info = sheet_data[category]
        existing_match = sheet_info["line_item_index"].get(normalize_name(counterparty))
        existing_heading = existing_match[0] if existing_match else None
        existing_subheading = existing_match[1] if existing_match else None
        existing_row = existing_match[2] if existing_match and len(existing_match) > 2 else None

        row_idx = file_rows.index[0]
        eff_text = st.session_state.get(f"override_{row_idx}_Effective Date", row.get("Effective Date"))
        term_text = st.session_state.get(f"override_{row_idx}_Term / Expiry", row.get("Term / Expiry"))
        value_text = st.session_state.get(f"override_{row_idx}_Contract Value", row.get("Contract Value"))
        eff_date = parse_date_flexible(eff_text)
        duration = parse_duration_months(term_text, eff_date)
        value = parse_amount(value_text, duration)
        raw_debug = {
            "Effective Date (raw)": eff_text,
            "Term / Expiry (raw)": term_text,
            "Contract Value (raw)": value_text,
        }

        # Ask the model to suggest where this fits among the sheet's real headings.
        suggested_heading, suggested_subheading = None, ""
        headings = sheet_info["headings"]
        cache_key = f"{file_name}::{category}"
        placement_cache = st.session_state.setdefault("placement_cache", {})
        if cache_key in placement_cache:
            suggested_heading, suggested_subheading = placement_cache[cache_key]
        elif headings and OPENAI_API_KEY and file_name in pdf_bytes_map:
            try:
                text = extract_pdf_text(io.BytesIO(pdf_bytes_map[file_name]))[:MAX_CHARS]
                prompt = build_placement_prompt(text, headings)
                raw = call_openai(prompt, OPENAI_API_KEY, OPENAI_MODEL)
                parsed = parse_json_response(raw)
                if "heading" in parsed:
                    suggested_heading = parsed.get("heading")
                    suggested_subheading = parsed.get("subheading", "") or ""
                placement_cache[cache_key] = (suggested_heading, suggested_subheading)
            except Exception:
                pass
        if suggested_heading not in [h["name"] for h in headings]:
            suggested_heading = headings[0]["name"] if headings else None

        plan.append({
            "file": file_name, "category": category, "counterparty": counterparty,
            "status": "update" if existing_row else "new",
            "sheet": category, "row_idx": row_idx,
            "effective_date": eff_date, "duration_months": duration, "contract_value": value,
            "raw_debug": raw_debug,
            "suggested_heading": suggested_heading, "suggested_subheading": suggested_subheading,
            "existing_heading": existing_heading,
            "existing_subheading": existing_subheading,
            "existing_row": existing_row,
        })

    if not plan:
        st.info("Nothing to plan yet — classify at least one contract into HC/CO/CS/GS/MM first.")
        return

    plan_by_sheet = {}
    grouped_plan = {}
    for item in plan:
        grouped_plan.setdefault(item["category"], []).append(item)

    for category in CATEGORY_ABBREVIATIONS:
        items = grouped_plan.get(category)
        if not items:
            continue
        category_label = f"{category} — {CLASSIFICATION_CATEGORIES.get(category, category)}  ({len(items)} contract{'s' if len(items) != 1 else ''})"
        with st.expander(category_label, expanded=True):
            for i, item in enumerate(items):
                if i > 0:
                    st.divider()
                st.markdown(f"**{item.get('counterparty', item['file'])}**")

                if item["status"] == "no_sheet":
                    st.warning(f"No sheet named '{item['category']}' in this workbook — can't place this contract.")
                    continue
                if item["status"] == "no_audit_data":
                    st.warning("No matching Contract Audit results found for this file.")
                    continue
                if item["status"] == "no_counterparty":
                    st.warning("Counterparty wasn't found in Contract Audit — fix that in Audit mode's 'Edit values' first.")
                    continue
                if item.get("existing_row"):
                    loc = (item.get("existing_heading") or "") + (
                        f" → {item['existing_subheading']}" if item.get("existing_subheading") else ""
                    )
                    st.info(
                        f"Already in the budget under **{loc or 'an existing row'}** — "
                        f"Apply will **overwrite** that row with the recalculated figures."
                    )

                # status == "new" or "update"
                row_idx = item["row_idx"]
                eff_date = item["effective_date"]
                duration = item["duration_months"]
                value = item["contract_value"]

                if eff_date is None:
                    entered = st.text_input(
                        "Effective Date wasn't found — enter it (e.g. '1 June 2026')",
                        key=f"fix_effdate_{item['file']}",
                    )
                    if entered.strip():
                        parsed_date = parse_date_flexible(entered)
                        if parsed_date is not None:
                            eff_date = parsed_date
                            st.session_state[f"override_{row_idx}_Effective Date"] = entered.strip()
                        else:
                            st.caption("Couldn't read that as a date — try a format like '1 June 2026'.")

                if duration is None:
                    entered = st.text_input(
                        "Term / Expiry wasn't found — enter it (e.g. '12 months')",
                        key=f"fix_term_{item['file']}",
                    )
                    if entered.strip():
                        parsed_duration = parse_duration_months(entered, eff_date)
                        if parsed_duration is not None:
                            duration = parsed_duration
                            st.session_state[f"override_{row_idx}_Term / Expiry"] = entered.strip()
                        else:
                            st.caption("Couldn't read a duration from that — try '12 months' or '2 years'.")

                # Recurring fees ("AED 42,000 per month") must be scaled by the
                # term. Re-parse after duration is known so a late term fix
                # cannot leave the monthly figure as if it were the total.
                raw_value_text = (item.get("raw_debug") or {}).get("Contract Value (raw)")
                if duration and raw_value_text:
                    restated = parse_amount(raw_value_text, duration)
                    if restated is not None:
                        value = restated

                if value is None:
                    entered = st.number_input(
                        "Contract Value wasn't found — enter the total contract value",
                        min_value=0.0, value=0.0, key=f"fix_value_{item['file']}",
                    )
                    if entered > 0:
                        value = entered
                        st.session_state[f"override_{row_idx}_Contract Value"] = str(entered)

                still_missing = []
                if eff_date is None:
                    still_missing.append("Effective Date")
                if duration is None:
                    still_missing.append("Term / Expiry")
                if value is None:
                    still_missing.append("Contract Value")
                if still_missing:
                    st.warning(f"Still missing: {', '.join(still_missing)} — fill in above, or enter the budgeted amount manually below.")
                    with st.expander("Why? Show what was actually read from Contract Audit"):
                        st.json(item.get("raw_debug", {}))
                        st.caption("If a field above shows a real value but still says 'missing', it's a parsing issue — otherwise the correction in Audit mode isn't reaching this contract yet.")

                headings = sheet_data[item["category"]]["headings"]
                heading_names = [h["name"] for h in headings]
                default_h_idx = heading_names.index(item["suggested_heading"]) if item["suggested_heading"] in heading_names else 0
                chosen_heading_name = st.selectbox("Heading", heading_names, index=default_h_idx, key=f"plan_heading_{item['file']}")
                chosen_heading = next(h for h in headings if h["name"] == chosen_heading_name)

                chosen_subheading_name = ""
                if chosen_heading["subheadings"]:
                    sub_names = [s["name"] for s in chosen_heading["subheadings"]]
                    default_s_idx = sub_names.index(item["suggested_subheading"]) if item["suggested_subheading"] in sub_names else 0
                    chosen_subheading_name = st.selectbox("Subheading", sub_names, index=default_s_idx, key=f"plan_sub_{item['file']}")

                account_names = st.session_state.get("confirmed_account_names") or []
                chosen_account_name = ""
                if account_names:
                    account_options = ["(none)"] + account_names
                    chosen_account_name = st.selectbox("Account Name (from Actuals Upload)", account_options, index=0, key=f"plan_account_{item['file']}")
                    chosen_account_name = "" if chosen_account_name == "(none)" else chosen_account_name

                name_candidates = [item["counterparty"], chosen_account_name]
                total, monthly_rate_calc, active_months, monthly_values, rf_meta = compute_reforecast(
                    value, eff_date, duration, year, actuals_lookup, name_candidates, through_month_idx,
                    budget_method=st.session_state.get("budget_method", "fixed"),
                )
                rf_meta = rf_meta or {}

                if not rf_meta.get("live", True):
                    st.info(
                        rf_meta.get("reason")
                        or "Not in existence at the end of the chosen month — no budget line this period."
                    )
                    continue

                orig_2026 = rf_meta.get("orig_2026_budget") or 0.0
                method = rf_meta.get("method", "fixed")
                locked_total = orig_2026 if method == "fixed" else (total or 0.0)
                default_total = round(locked_total, 2) if locked_total is not None else 0.0
                amount_key = f"plan_total_{item['file']}_{method}_{default_total}"
                reviewer_total = st.number_input(
                    f"Budgeted amount for {year} (locked 2026 envelope)" if method == "fixed" else f"Budgeted amount for {year} (pro-rated)",
                    value=default_total,
                    key=amount_key,
                    disabled=method == "fixed",
                )
                if method == "fixed":
                    reviewer_total = default_total

                raw_dbg = item.get("raw_debug") or {}
                monthly_orig = rf_meta.get("orig_monthly_rate") or 0.0
                periods_used = rf_meta.get("periods")
                if periods_used is None and monthly_orig:
                    periods_used = round(orig_2026 / monthly_orig) if monthly_orig else 0
                v_start = rf_meta.get("validity_start") or "?"
                v_end = rf_meta.get("validity_end") or "?"
                start_txt = (
                    pd.Timestamp(eff_date).strftime("%d %b %Y")
                    if eff_date is not None else str(raw_dbg.get("Effective Date (raw)") or "not found")
                )
                term_txt = raw_dbg.get("Term / Expiry (raw)") or (f"{duration} months" if duration else "not found")
                value_txt = raw_dbg.get("Contract Value (raw)") or (f"{value:,.0f}" if value else "not found")
                value_num = value if value else 0.0
                dur_n = int(duration) if duration else 0
                per_n = int(periods_used) if periods_used is not None else 0
                with st.expander("How this envelope was calculated", expanded=True):
                    st.markdown(
                        f"""
- **Extracted value:** {value_txt}  
- **Term:** {term_txt} → **{dur_n or "?"} months**  
- **Start date:** {start_txt}  
- **Monthly rate** = contract figure ÷ term months  
  `{value_num:,.0f} ÷ {dur_n or "?"} = {monthly_orig:,.2f}`  
- **{year} validity:** {v_start}–{v_end} = **{per_n or "?"} month(s)**  
  (start month is included; the month before start is included only if that month has an actual)  
- **Column B envelope** = monthly rate × {year} validity months  
  `{monthly_orig:,.2f} × {per_n or "?"} = {orig_2026:,.2f}`
                        """.strip()
                    )

                active_months = active_months or []
                monthly_values = monthly_values or {}
                if method != "fixed" and monthly_values and total and abs(reviewer_total - total) > 0.01:
                    past = {m: v for m, v in monthly_values.items() if m <= through_month_idx}
                    future = {m: v for m, v in monthly_values.items() if m > through_month_idx}
                    leftover = reviewer_total - sum(past.values())
                    if future:
                        each = leftover / len(future)
                        monthly_values = {**past, **{m: each for m in future}}
                        monthly_rate_calc = each

                monthly_values = {m: max(float(v), 0.0) for m, v in monthly_values.items()}
                recorded = lookup_actual_amount(actuals_lookup, name_candidates, through_month_idx) if actuals_lookup else None
                n_act = rf_meta.get("actual_months", 0)
                rem = rf_meta.get("remaining_period", 0)
                cum = rf_meta.get("cumulative_actual", 0.0)
                if rf_meta.get("insufficient"):
                    over = cum - (orig_2026 if method == "fixed" else (value or 0.0))
                    st.error(
                        f"Budget insufficient for **{item['counterparty']}**. "
                        f"Actuals to {MONTH_NAMES[through_month_idx-1]} are {cum:,.2f} versus "
                        f"{'2026 budget ' + format(orig_2026, ',.0f') if method == 'fixed' else 'contract value ' + format(value or 0, ',.0f')}. "
                        f"Overspent by {over:,.2f}. Remaining months are set to 0 — no negative figures."
                    )
                if method == "fixed":
                    act_end = MONTH_NAMES[through_month_idx - 1]
                    expiry_m = int(rf_meta.get("expiry_month") or 12)
                    expiry_label = MONTH_NAMES[expiry_m - 1] if 1 <= expiry_m <= 12 else "Dec"
                    if rem <= 0:
                        st.caption(
                            f"Actual period Jan–{act_end}. No forecast period left "
                            f"(contract expires {expiry_label}). "
                            f"Column B stays {orig_2026:,.0f}. YTD actuals {cum:,.0f}. "
                            f"Months after expiry are 0."
                        )
                    else:
                        fc_start = MONTH_NAMES[through_month_idx]
                        st.caption(
                            f"Actual period Jan–{act_end} (posted actuals only). "
                            f"Forecast period {fc_start}–{expiry_label} "
                            f"({rem} month{'s' if rem != 1 else ''}; "
                            f"min(12, expiry {expiry_m}) − {through_month_idx}). "
                            f"Column B = {orig_2026:,.0f}. "
                            f"Forecast rate = ({orig_2026:,.0f} − {cum:,.0f}) / {rem} "
                            f"= {monthly_rate_calc:,.2f}/month. "
                            f"Months after {expiry_label} are 0."
                            + (f" {act_end} actual = {recorded:,.2f}." if recorded is not None else f" No {act_end} actual posted.")
                        )
                elif recorded is not None:
                    st.caption(
                        f"{MONTH_NAMES[through_month_idx-1]} actual = {recorded:,.2f}. "
                        f"Flexible rate = ({value:,.0f} − {cum:,.0f}) / ({int(duration)} − {n_act}) "
                        f"= {monthly_rate_calc:,.2f}/month."
                    )
                elif monthly_rate_calc:
                    st.caption(
                        f"No {MONTH_NAMES[through_month_idx-1]} actual posted. "
                        f"Original 2026 monthly rate = {rf_meta.get('orig_monthly_rate', monthly_rate_calc):,.2f}."
                    )
                else:
                    st.caption("Could not compute a monthly rate — check Effective Date, Term, and Contract Value.")

                plan_by_sheet.setdefault(item["category"], []).append({
                    "counterparty": item["counterparty"],
                    "insertion_row": find_insertion_row(chosen_heading, chosen_subheading_name),
                    "existing_row": item.get("existing_row"),
                    "prorated_total": reviewer_total,
                    "monthly_rate": monthly_rate_calc or 0.0,
                    "active_months": active_months or list(range(1, 13)),
                    "monthly_values": monthly_values,
                    "account_name": chosen_account_name,
                    "orig_2026_budget": rf_meta.get("orig_2026_budget", 0.0),
                    "orig_monthly_rate": rf_meta.get("orig_monthly_rate", 0.0),
                    "ready_from": rf_meta.get("ready_from", 1),
                })

                preview_vals = [monthly_values.get(m, 0.0) for m in range(1, 13)]
                preview_df = pd.DataFrame(
                    [[item["counterparty"], reviewer_total] + preview_vals],
                    columns=["Line Item", "Budget"] + MONTH_NAMES,
                )
                st.markdown(preview_df.to_html(index=False, escape=False, na_rep="0"), unsafe_allow_html=True)
                if actuals_lookup:
                    found = []
                    for m in range(1, 13):
                        hit = lookup_actual_amount(actuals_lookup, name_candidates, m)
                        if hit is not None:
                            found.append(f"{MONTH_NAMES[m-1]} {hit:,.0f}")
                    if found:
                        st.caption("Actuals matched to this vendor: " + " · ".join(found))
                    else:
                        st.caption("No actuals rows matched this vendor name. Check Account Name in Step 3 / the dropdown above.")
                    sel_hit = lookup_actual_amount(actuals_lookup, name_candidates, through_month_idx)
                    if sel_hit is None:
                        st.warning(
                            f"No {MONTH_NAMES[through_month_idx-1]} actual matched for "
                            f"**{item['counterparty']}**. That month is stored as 0. "
                            f"If the figure sits under a different account name, pick it in "
                            f"**Account Name (from Actuals Upload)**."
                        )

    st.markdown('<div class="step-label">3 · Apply</div>', unsafe_allow_html=True)
    if st.button("Apply to workbook", type="primary", disabled=not plan_by_sheet):
        workbook_bytes = write_forecast_workbook(wb, plan_by_sheet)
        st.session_state.forecast_workbook_bytes = workbook_bytes
        st.session_state.forecast_sheets_touched = list(plan_by_sheet.keys())
        st.session_state.max_unlocked_step = max(st.session_state.max_unlocked_step, 5)
        baseline = {}
        for items in plan_by_sheet.values():
            for ins in items:
                rate = ins.get("orig_monthly_rate") or 0.0
                ready = ins.get("ready_from") or 1
                orig_months = {m: (rate if m >= ready else 0.0) for m in range(1, 13)}
                baseline[normalize_name(ins["counterparty"])] = {
                    "orig_2026": ins.get("orig_2026_budget") or 0.0,
                    "orig_monthly_rate": rate,
                    "orig_months": orig_months,
                    "display_name": ins["counterparty"],
                }
        st.session_state.fixed_budget_baseline = baseline

    if st.session_state.get("forecast_workbook_bytes"):
        st.markdown('<div class="step-label">Updated sheets</div>', unsafe_allow_html=True)
        result_wb = openpyxl.load_workbook(io.BytesIO(st.session_state.forecast_workbook_bytes), data_only=False)
        for sheet_name in st.session_state.forecast_sheets_touched:
            ws = result_wb[sheet_name]
            st.caption(f"**{sheet_name}**")
            rows_out = []
            for row in range(3, ws.max_row + 1):
                name = ws.cell(row=row, column=1).value
                if name is None:
                    continue
                budget = ws.cell(row=row, column=2).value
                month_vals = [ws.cell(row=row, column=4 + m).value for m in range(12)]
                rows_out.append([name, budget] + month_vals)
            preview_df = pd.DataFrame(rows_out, columns=["Line Item", "Budget"] + MONTH_NAMES)
            st.markdown(preview_df.to_html(index=False, escape=False, na_rep=""), unsafe_allow_html=True)

        st.download_button(
            "Download updated workbook",
            data=st.session_state.forecast_workbook_bytes,
            file_name=f"budget_updated_{year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption("Heading, subheading, and grand-total cells now contain live Excel SUM formulas, so the workbook stays fully editable.")

def render_variance_mode():
    st.caption("Compares actual spend against your finalized budget by cohort and by line item — pure arithmetic, no AI, no API calls.")

    if not st.session_state.get("forecast_workbook_bytes"):
        st.info("Complete **Step 4 (Forecasting)** and click 'Apply to workbook' first — variance tracking compares actuals against that finalized budget.")
        return

    dump_mapping = st.session_state.get("actuals_dump_mapping")
    if not dump_mapping:
        st.info("Upload an actuals dump in **Step 3** first — variance tracking needs real spend data to compare against the budget.")
        return

    df_dump, date_col, dump_name_col, amount_col, account_col = dump_mapping
    match_col = account_col or dump_name_col

    c1, c2 = st.columns(2)
    year = c1.number_input("Calendar year", min_value=2020, max_value=2100, value=2026, step=1, key="variance_year")
    through_month = c2.selectbox("Actuals known through", MONTH_NAMES, index=11, key="variance_through")
    through_idx = MONTH_NAMES.index(through_month) + 1

    wb = openpyxl.load_workbook(io.BytesIO(st.session_state.forecast_workbook_bytes), data_only=False)

    # Read budgeted figures straight from the leaf line-item cells (not the
    # heading/subheading/total formulas) so nothing depends on Excel having
    # actually recalculated the workbook yet.
    line_item_to_sheet = {}
    sheet_budget = {}
    line_item_budget = {}  # (sheet, normalized_name) -> {"annual": x, "to_date": y, "display_name": ...}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        nodes = build_sheet_structure(ws)
        annual_total, to_date_total = 0.0, 0.0
        for node in nodes:
            if node["type"] != "line_item":
                continue
            row = node["row"]
            name_norm = normalize_name(node["name"])
            line_item_to_sheet[name_norm] = sheet_name
            annual = 0.0
            to_date = 0.0
            for m in range(1, 13):
                val = ws.cell(row=row, column=3 + m).value
                val = val if isinstance(val, (int, float)) else 0.0
                annual += val
                if m <= through_idx:
                    to_date += val
            line_item_budget[(sheet_name, name_norm)] = {
                "annual": annual, "to_date": to_date, "display_name": node["name"],
            }
            annual_total += annual
            to_date_total += to_date
        sheet_budget[sheet_name] = {"annual": annual_total, "to_date": to_date_total}

    # Aggregate actuals through the chosen month, matched by name.
    df = df_dump.copy()
    df[date_col] = pd.to_datetime(df[date_col], errors="coerce")
    df = df[(df[date_col].dt.year == year) & (df[date_col].dt.month <= through_idx)]
    df["_name_norm"] = df[match_col].apply(normalize_name)
    try:
        df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    except Exception:
        pass
    actual_by_name = df.groupby("_name_norm")[amount_col].sum().to_dict()

    sheet_actual = {}
    line_item_actual = {}
    for name_norm, amt in actual_by_name.items():
        sheet_name = line_item_to_sheet.get(name_norm)
        if sheet_name is None:
            continue
        sheet_actual[sheet_name] = sheet_actual.get(sheet_name, 0.0) + amt
        line_item_actual[(sheet_name, name_norm)] = amt

    def status_for(actual, budget):
        if not budget:
            return ("No budget", "gray") if not actual else ("Spend with no budget", "red")
        pct = actual / budget * 100
        if pct > 115:
            return (f"Over pace ({pct:.0f}% of prorated budget)", "red")
        if pct < 80:
            return (f"Under pace ({pct:.0f}% of prorated budget)", "orange")
        return (f"On track ({pct:.0f}% of prorated budget)", "green")

    st.markdown('<div class="step-label">Cohort summary</div>', unsafe_allow_html=True)
    for sheet_name in wb.sheetnames:
        budget_to_date = sheet_budget.get(sheet_name, {}).get("to_date", 0.0)
        annual_budget = sheet_budget.get(sheet_name, {}).get("annual", 0.0)
        actual = sheet_actual.get(sheet_name, 0.0)
        label, color = status_for(actual, budget_to_date)
        full_name = CLASSIFICATION_CATEGORIES.get(sheet_name, sheet_name)

        with st.container(border=True):
            cols = st.columns([2, 1.3, 1.3, 1.3, 2])
            cols[0].markdown(f"**{sheet_name} — {full_name}**")
            cols[1].metric("Annual budget", f"{annual_budget:,.0f}")
            cols[2].metric(f"Budget to {through_month}", f"{budget_to_date:,.0f}")
            cols[3].metric(f"Actual to {through_month}", f"{actual:,.0f}")
            variance = actual - budget_to_date
            sign = "+" if variance > 0 else ""
            cols[4].markdown(f":{color}[**{label}**]")
            cols[4].caption(f"Variance: {sign}{variance:,.0f}")

    st.markdown('<div class="step-label">Month flags vs original 2026 plan</div>', unsafe_allow_html=True)
    st.caption("Compares the chosen month's actual with the original monthly budget (not the reforecasted figure).")
    baseline = st.session_state.get("fixed_budget_baseline") or {}
    monthly_lookup = build_monthly_actuals_lookup(df_dump, date_col, match_col, amount_col, year)
    if account_col and account_col != dump_name_col:
        extra = build_monthly_actuals_lookup(df_dump, date_col, dump_name_col, amount_col, year)
        for k, v in extra.items():
            monthly_lookup.setdefault(k, v)
    month_flagged = False
    names_to_check = set(line_item_to_sheet.keys()) | set(baseline.keys())
    for name_norm in sorted(names_to_check):
        actual_m = lookup_actual_amount(monthly_lookup, [name_norm], through_idx)
        if actual_m is None:
            continue
        orig_rate = (baseline.get(name_norm) or {}).get("orig_monthly_rate")
        display = (baseline.get(name_norm) or {}).get("display_name") or name_norm
        if not orig_rate:
            info = next((v for (s, n), v in line_item_budget.items() if n == name_norm), None)
            if info:
                live_months = 12
                orig_rate = (info["annual"] / live_months) if info["annual"] else 0
                display = info["display_name"]
        if not orig_rate:
            continue
        pct = actual_m / orig_rate * 100
        if pct > 150:
            month_flagged = True
            st.error(
                f"**{display}** — {MONTH_NAMES[through_idx-1]} actual {actual_m:,.0f} is "
                f"{pct:.0f}% of the original monthly budget {orig_rate:,.0f} "
                f"(about {pct/100:.1f}× planned)."
            )
        elif pct > 115:
            month_flagged = True
            st.warning(
                f"**{display}** — {MONTH_NAMES[through_idx-1]} actual {actual_m:,.0f} is "
                f"{pct:.0f}% of the original monthly budget {orig_rate:,.0f}."
            )
    if not month_flagged:
        st.success(f"No line is far over its original {through_month} budget.")

    st.markdown('<div class="step-label">Line items to review</div>', unsafe_allow_html=True)
    st.caption("Flagged when a line's actual spend to date is well outside its own prorated budget — possible candidates for reallocating budget between lines.")

    flagged_any = False
    for (sheet_name, name_norm), info in line_item_budget.items():
        actual = line_item_actual.get((sheet_name, name_norm), 0.0)
        budget_to_date = info["to_date"]
        if budget_to_date == 0 and actual == 0:
            continue
        if budget_to_date == 0:
            pct = None
        else:
            pct = actual / budget_to_date * 100
        if pct is not None and 80 <= pct <= 115:
            continue  # within normal range, don't clutter the list
        flagged_any = True
        full_name = CLASSIFICATION_CATEGORIES.get(sheet_name, sheet_name)
        if pct is None:
            msg = f"**{info['display_name']}** ({sheet_name}) — {actual:,.0f} spent with no budget allocated to {through_month}."
        elif pct > 115:
            msg = f"**{info['display_name']}** ({sheet_name}) — {pct:.0f}% of its prorated budget used ({actual:,.0f} vs {budget_to_date:,.0f} expected). Consider reallocating from an underspent line in {full_name}."
        else:
            msg = f"**{info['display_name']}** ({sheet_name}) — only {pct:.0f}% of its prorated budget used ({actual:,.0f} vs {budget_to_date:,.0f} expected). Budget here may be available to reallocate."
        st.warning(msg)

    if not flagged_any:
        st.success("No line items are significantly off pace — everything falls within a normal range of its prorated budget.")
# ---------------------------------------------------------------------------
# Wizard navigation — each step must be confirmed before the next unlocks
# ---------------------------------------------------------------------------
STEP_NAMES = {1: "1. Audit", 2: "2. Classify", 3: "3. Actuals", 4: "4. Forecast", 5: "5. Variance"}
nav_cols = st.columns(5)
for step_num, label in STEP_NAMES.items():
    with nav_cols[step_num - 1]:
        is_current = st.session_state.wizard_step == step_num
        unlocked = step_num <= st.session_state.max_unlocked_step
        if st.button(
            ("▶ " if is_current else "") + label,
            key=f"nav_step_{step_num}",
            disabled=not unlocked,
            use_container_width=True,
            type="primary" if is_current else "secondary",
        ):
            st.session_state.wizard_step = step_num
            st.rerun()

st.markdown("<div style='height: 0.5rem'></div>", unsafe_allow_html=True)

if st.session_state.wizard_step == 1:
    render_audit_mode()
elif st.session_state.wizard_step == 2:
    render_classification_mode()
elif st.session_state.wizard_step == 3:
    render_actuals_upload_mode()
elif st.session_state.wizard_step == 4:
    render_forecast_mode()
else:
    render_variance_mode()

save_persisted_state()
